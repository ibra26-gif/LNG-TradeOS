#!/usr/bin/env python3
"""
fetch_gassco_umm.py — Gassco UMM + nominations scraper

Primary source: XLSX export from https://umm.gassco.no/xlexport
  - Clean structured data: Asset, Status, Type, Start, Stop, Technical, Available
  - Impact = Technical - Available (mcm/d)
  - Published once per gasday, includes all active + recent events

Fallback: Atom feeds
  - https://umm.gassco.no/atom.xml         (all UMMs)
  - https://umm.gassco.no/realTimeAtom.xml (real-time)

Outputs TWO files:
  - /data/umm_active.json         (banner on dashboard)
  - /data/gassco_nominations.json (Norway "D" column — aggregated total)
    Note: Gassco publishes per-terminal on the webpage but not in XLSX.
    This scraper parses the homepage HTML for real-time terminal values.
"""

import json
import os
import re
import sys
from datetime import datetime, timezone, timedelta
from io import BytesIO

import requests

try:
    import openpyxl
    HAVE_XLSX = True
except ImportError:
    HAVE_XLSX = False
    print("WARN: openpyxl not installed, XLSX parsing unavailable")

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}

TIMEOUT = 30

XLSX_URL = "https://umm.gassco.no/xlexport"
ATOM_URL = "https://umm.gassco.no/atom.xml"
REALTIME_ATOM_URL = "https://umm.gassco.no/realTimeAtom.xml"
HOMEPAGE_URL = "https://umm.gassco.no/"


# =============================================================================
# UMM — from XLSX (primary)
# =============================================================================

def fetch_umms_from_xlsx(session):
    """Download xlexport, parse active events. Returns list of UMM dicts.

    Note: xlexport endpoint may require session cookies from homepage visit first.
    We prime the session by fetching the homepage, then request XLSX.
    """
    if not HAVE_XLSX:
        return None
    try:
        # Prime session: fetch homepage first to establish cookies
        print(f"Priming session with {HOMEPAGE_URL}...")
        prime = session.get(HOMEPAGE_URL, timeout=TIMEOUT)
        print(f"  Homepage: {prime.status_code}, cookies now: {len(session.cookies)}")

        # Now fetch XLSX with browser-like headers + referer
        print(f"Downloading {XLSX_URL}...")
        xlsx_headers = {
            "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/vnd.ms-excel,*/*",
            "Referer": HOMEPAGE_URL,
            "Sec-Fetch-Dest": "document",
            "Sec-Fetch-Mode": "navigate",
            "Sec-Fetch-Site": "same-origin",
        }
        r = session.get(XLSX_URL, timeout=TIMEOUT, headers=xlsx_headers)
        if r.status_code != 200:
            print(f"  XLSX status {r.status_code}")
            return None

        # Sanity-check content before trying to parse
        content = r.content
        ct = (r.headers.get("Content-Type") or "").lower()
        print(f"  Got {len(content)} bytes, content-type: {ct}")

        # XLSX files start with PK (zip magic)
        if not content[:2] == b"PK":
            # Server returned HTML error page instead of XLSX - log detailed info for debugging
            preview = content[:500].decode("utf-8", errors="replace")
            print(f"  Not a valid XLSX. First 500 chars:")
            print(f"  {preview!r}")
            # Also log tail for stack-trace-style errors
            if len(content) > 1000:
                tail = content[-500:].decode("utf-8", errors="replace")
                print(f"  Last 500 chars:")
                print(f"  {tail!r}")
            # And log request headers sent
            print(f"  Request was sent with headers: {dict(session.headers)}")
            return None

        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
        ws = wb.active  # "Past Events" sheet
        print(f"  Loaded {ws.max_row} rows from sheet '{ws.title}'")

        # Row 2 is main header, row 3 is sub-header (Start/Stop under "Event", Technical/Available under "Capacity")
        # Data starts row 4. Columns: A=MsgID, B=Asset, C=Status, D=Type, E=EventType,
        # F=PubDate, G=EventStart, H=EventStop, I=Unit, J=Technical, K=Available
        umms = []
        for r in range(4, ws.max_row + 1):
            msg_id = ws.cell(r, 1).value
            asset = ws.cell(r, 2).value
            status = ws.cell(r, 3).value
            unavail_type = ws.cell(r, 4).value
            event_type = ws.cell(r, 5).value
            pub_date = ws.cell(r, 6).value
            event_start = ws.cell(r, 7).value
            event_stop = ws.cell(r, 8).value
            unit = ws.cell(r, 9).value
            technical = ws.cell(r, 10).value
            available = ws.cell(r, 11).value

            if not msg_id or not asset:
                continue

            # Compute impact (negative = reduction)
            impact_mcm = None
            if technical is not None and available is not None:
                try:
                    impact_mcm = float(available) - float(technical)
                except (TypeError, ValueError):
                    pass

            umms.append({
                "msg_id": str(msg_id).strip(),
                "asset": str(asset).strip(),
                "status": str(status).strip() if status else "",
                "unavail_type": str(unavail_type).strip() if unavail_type else "",
                "event_type": str(event_type).strip() if event_type else "",
                "published": pub_date.isoformat() if isinstance(pub_date, datetime) else str(pub_date or ""),
                "event_start": event_start.isoformat() if isinstance(event_start, datetime) else str(event_start or ""),
                "event_stop": event_stop.isoformat() if isinstance(event_stop, datetime) else str(event_stop or ""),
                "unit": str(unit).strip() if unit else "mcm/d",
                "technical": float(technical) if technical is not None else None,
                "available": float(available) if available is not None else None,
                "impact_mcm": impact_mcm,
            })
        print(f"  Parsed {len(umms)} UMM entries")
        return umms
    except Exception as e:
        print(f"  XLSX fetch failed: {type(e).__name__}: {e}")
        return None


def filter_active(umms, now=None):
    """Active = status='Active' AND event_start <= now <= event_stop (or status=Active with no dates)."""
    if now is None:
        now = datetime.now(timezone.utc)
    active = []
    for u in umms:
        if u["status"].lower() != "active":
            continue
        # Parse dates
        try:
            start = datetime.fromisoformat(u["event_start"].replace("Z", "+00:00")) if u["event_start"] else None
            stop = datetime.fromisoformat(u["event_stop"].replace("Z", "+00:00")) if u["event_stop"] else None
            # Make timezone-aware
            if start and start.tzinfo is None:
                start = start.replace(tzinfo=timezone.utc)
            if stop and stop.tzinfo is None:
                stop = stop.replace(tzinfo=timezone.utc)
        except (ValueError, AttributeError):
            start = stop = None

        # Active if now is between start and stop
        if start and stop:
            if start <= now <= stop:
                active.append(u)
        elif start and not stop:
            if start <= now:
                active.append(u)
        # No dates: assume active since status says so
        else:
            active.append(u)
    return active


def format_for_dashboard(active_umms, limit=5):
    """Shape for banner display — keep top N by impact magnitude."""
    # Sort by impact (biggest reduction first)
    ranked = sorted(
        active_umms,
        key=lambda u: abs(u["impact_mcm"] or 0),
        reverse=True
    )
    return [{
        "title": f"{u['event_type']} · {u['unavail_type']}" if u['event_type'] else u['unavail_type'],
        "asset": u["asset"],
        "impact": (f"{u['impact_mcm']:+.1f} mcm/d" if u["impact_mcm"] is not None else ""),
        "from": u["event_start"][:10] if u["event_start"] else "",
        "to": u["event_stop"][:10] if u["event_stop"] else "",
        "type": u["unavail_type"].lower() if u["unavail_type"] else "",
        "published": u["published"],
        "msg_id": u["msg_id"],
    } for u in ranked[:limit]]


# =============================================================================
# Nominations — from homepage HTML scrape
# =============================================================================

def fetch_nominations_from_homepage(session):
    """Scrape the 'REAL TIME INFORMATION' block from umm.gassco.no homepage.
    Returns dict with per-terminal values and total.

    HTML structure confirmed from live inspection (2026-04-18):
      Each terminal is rendered as a block containing the terminal name
      followed by <div class="value">NUMBER</div> with MSm3 unit nearby.
    """
    try:
        print(f"Fetching {HOMEPAGE_URL} for nominations...")
        r = session.get(HOMEPAGE_URL, timeout=TIMEOUT)
        if r.status_code != 200:
            print(f"  Homepage status {r.status_code}")
            return None
        html = r.text
        print(f"  Got {len(html)} chars of HTML")

        # Diagnostic: does the HTML contain expected markers?
        markers = {
            'REAL TIME': 'REAL TIME' in html,
            'Dornum': 'Dornum' in html,
            'class="value"': 'class="value"' in html,
            "class='value'": "class='value'" in html,
            'MSm': 'MSm' in html,
            '67.5': '67.5' in html,  # today's Dornum value - won't match tomorrow
        }
        print(f"  HTML markers present: {markers}")

        # If key markers are missing, dump a sample so we can see what came back
        if not markers['Dornum'] or not markers['REAL TIME']:
            print(f"  UNEXPECTED HTML STRUCTURE. First 800 chars:")
            print(f"  {html[:800]!r}")
            if len(html) > 2000:
                mid = len(html) // 2
                print(f"  Middle 800 chars (from pos {mid}):")
                print(f"  {html[mid:mid+800]!r}")

        # Extract gasday date
        date_m = re.search(r"gasday\s+(\d{4}-\d{2}-\d{2})", html, flags=re.IGNORECASE)
        gasday = date_m.group(1) if date_m else datetime.now(timezone.utc).strftime("%Y-%m-%d")

        # Known Gassco terminals (order matters - match longer names first to avoid
        # "St.Fergus" being partially matched by "St Fergus" and vice versa)
        known_terminals = [
            "Field Deliveries into SEGAL",
            "Aggregated other exit points",
            "Sum exit nominations NCS",
            "St.Fergus", "St Fergus",
            "Dornum", "Emden", "Nybro", "Dunkerque", "Zeebrugge", "Easington",
        ]

        by_terminal = {}

        # Strategy 1: Find the REAL TIME INFORMATION section boundary
        rt_start = html.find("REAL TIME")
        rt_end = html.find("FILTERS", rt_start) if rt_start > 0 else -1
        if rt_end < 0:
            rt_end = html.find("EVENTS", rt_start) if rt_start > 0 else -1
        if rt_end < 0:
            rt_end = min(rt_start + 10000, len(html)) if rt_start > 0 else len(html)
        block = html[rt_start:rt_end] if rt_start > 0 else html
        print(f"  REAL TIME block: {len(block)} chars (from pos {rt_start} to {rt_end})")

        # Count value divs in the block to verify structure
        value_count = len(re.findall(r'class=["\']value["\']', block))
        print(f"  Found {value_count} class='value' divs in block")

        # Strategy 2: For each known terminal, search forward for the next
        # <div class="value">NUMBER</div> within a reasonable window
        for t in known_terminals:
            idx = block.find(t)
            if idx == -1:
                continue
            # Look ahead up to 500 chars for the value div
            window = block[idx:idx+500]
            # Primary pattern: <div class="value">NUMBER</div>
            val_m = re.search(r'class=["\']value["\']>\s*(-?\d+(?:\.\d+)?)\s*<', window)
            if val_m:
                by_terminal[t] = float(val_m.group(1))
                continue
            # Fallback pattern: any number followed by MSm/mcm context
            val_m = re.search(r'(-?\d+(?:\.\d+)?)\s*(?:<[^>]*>\s*)*(?:MSm|mcm|m\s*³|m\s*3)', window, flags=re.IGNORECASE)
            if val_m:
                by_terminal[t] = float(val_m.group(1))

        # Strategy 3: If very few terminals matched, extract all div.value numbers
        # in order and map to known terminal sequence on the page
        if len(by_terminal) < 5:
            all_values = re.findall(r'class=["\']value["\']>\s*(-?\d+(?:\.\d+)?)\s*<', block)
            print(f"  Strategy 3 fallback: {len(all_values)} values extracted in order: {all_values[:12]}")
            page_order = [
                "Dornum", "Emden", "Nybro", "Dunkerque", "Zeebrugge", "Easington",
                "St.Fergus", "Field Deliveries into SEGAL",
                "Aggregated other exit points", "Sum exit nominations NCS",
            ]
            if len(all_values) >= 5:
                by_terminal = {}
                for i, name in enumerate(page_order):
                    if i < len(all_values):
                        try:
                            by_terminal[name] = float(all_values[i])
                        except ValueError:
                            pass

        # "Sum exit nominations NCS" is the authoritative total if present
        # Otherwise sum the entry terminals (exclude aggregates that would double-count)
        ENTRY_POINTS = {"Dornum", "Emden", "Nybro", "Dunkerque", "Zeebrugge",
                        "Easington", "St.Fergus", "St Fergus",
                        "Field Deliveries into SEGAL"}
        if "Sum exit nominations NCS" in by_terminal:
            total = by_terminal["Sum exit nominations NCS"]
        else:
            total = sum(v for k, v in by_terminal.items() if k in ENTRY_POINTS)

        print(f"  Found {len(by_terminal)} terminals, total {total:.1f} mcm/d")
        for k, v in by_terminal.items():
            print(f"    {k:40s} = {v:6.1f}")

        return {
            "gasday": gasday,
            "by_terminal": by_terminal,
            "today_total_mcm": total if total > 0 else None,
        } if by_terminal else None

    except Exception as e:
        print(f"  Homepage nomination scrape failed: {type(e).__name__}: {e}")
        import traceback; traceback.print_exc()
        return None


# =============================================================================
# Main
# =============================================================================

def main():
    session = requests.Session()
    session.headers.update(HEADERS)

    # ---- UMMs from XLSX ----
    umms = fetch_umms_from_xlsx(session)
    active_umms = filter_active(umms) if umms else []
    display_umms = format_for_dashboard(active_umms) if active_umms else []

    umm_result = {
        "active": display_umms,
        "_meta": {
            "updated": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
            "source": "gassco xlexport",
            "total_messages": len(umms) if umms else 0,
            "active_count": len(active_umms),
        }
    }

    out_dir = os.environ.get("OUTPUT_DIR", "data")
    os.makedirs(out_dir, exist_ok=True)
    umm_path = os.path.join(out_dir, "umm_active.json")
    with open(umm_path, "w") as f:
        json.dump(umm_result, f, indent=2, ensure_ascii=False, default=str)
    print(f"\nDONE {umm_path} — {len(active_umms)} active UMMs")

    # ---- Nominations from homepage HTML ----
    noms = fetch_nominations_from_homepage(session)
    nom_result = {
        "gasday": noms["gasday"] if noms else "",
        "by_terminal": noms["by_terminal"] if noms else {},
        "today_total_mcm": noms["today_total_mcm"] if noms else None,
        "_meta": {
            "updated": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
            "source": "gassco homepage scrape",
        }
    }
    nom_path = os.path.join(out_dir, "gassco_nominations.json")
    with open(nom_path, "w") as f:
        json.dump(nom_result, f, indent=2, ensure_ascii=False)
    print(f"DONE {nom_path} — total {nom_result['today_total_mcm']} mcm/d")


if __name__ == "__main__":
    main()
