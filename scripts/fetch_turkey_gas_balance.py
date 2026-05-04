#!/usr/bin/env python3
"""Build Turkey gas balance from EPDK monthly natural-gas reports.

Balance rule:
  LNG imports + domestic production + pipeline imports = sectorial gas demand

The numbers come from EPDK only:
  - PDF Table 2.1.2: LNG vs pipeline import split.
  - Excel appendix Table 1 or DOCX summary table: domestic production.
  - Excel appendix Table 4: import country rows.
  - Excel appendix Table 6 / DEPOLAMA: storage stock.
  - Excel appendix Table 7: sectorial gas demand.

Power generation by source is separate from EPDK. It is wired to TEİAŞ
monthly electricity production-consumption workbooks. Hydro, coal, wind,
solar, and alternative fuels come from the workbook rows only.
No fallback rows are estimated.
"""

from __future__ import annotations

import html
import json
import re
import tempfile
import unicodedata
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import urljoin
from urllib.request import Request, urlopen
from zipfile import ZipFile
from xml.etree import ElementTree as ET


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "data" / "turkey_gas_balance.json"

EPDK_PAGE = "https://www.epdk.gov.tr/Detay/Icerik/3-0-95-1007/dogal-gazaylik-sektor-raporu"
TEIAS_POWER_PAGE = "https://www.teias.gov.tr/aylik-elektrik-uretim-tuketim-raporlari"
TEIAS_GALLERY_API = "https://www.teias.gov.tr/api/gallery?locale=tr-TR&slug="
TEIAS_FILE_BASE = "https://webim.teias.gov.tr/file/"
TEIAS_FIXED_HISTORY_START_YEAR = 2022
TEIAS_FIXED_HISTORY_END_YEAR = 2025

MONTHS_TR = {
    "ocak": "01",
    "subat": "02",
    "mart": "03",
    "nisan": "04",
    "mayis": "05",
    "haziran": "06",
    "temmuz": "07",
    "agustos": "08",
    "eylul": "09",
    "ekim": "10",
    "kasim": "11",
    "aralik": "12",
}

MONTHS_EN = {
    "january": "01",
    "february": "02",
    "march": "03",
    "april": "04",
    "may": "05",
    "june": "06",
    "july": "07",
    "august": "08",
    "september": "09",
    "october": "10",
    "november": "11",
    "december": "12",
}


def source_map(gas_status: str, power_status: str, error: str | None = None) -> list[dict[str, Any]]:
    return [
        {
            "item": "LNG imports and pipeline imports",
            "source": "EPDK monthly natural gas sector report PDF",
            "sourceUrl": EPDK_PAGE,
            "pull": "PDF Table 2.1.2 / Gazın Türüne Göre Dağılım: LNG and Boru Gazı import split.",
            "status": gas_status,
        },
        {
            "item": "Domestic production",
            "source": "EPDK monthly report Excel appendix",
            "sourceUrl": EPDK_PAGE,
            "pull": "Appendix Table 1, Genel Toplam production row; older 2025 DOCX reports use the overview balance table.",
            "status": gas_status,
        },
        {
            "item": "Sectorial gas demand",
            "source": "EPDK monthly report Excel appendix",
            "sourceUrl": EPDK_PAGE,
            "pull": "Appendix Table 7. Power generation/conversion = section 1 total; industry = section 4; residential = Konut; sectorial demand total = Genel Toplam.",
            "status": gas_status,
        },
        {
            "item": "Storage",
            "source": "EPDK monthly report Excel appendix",
            "sourceUrl": EPDK_PAGE,
            "pull": "Appendix Table 6 / DEPOLAMA. Underground, LNG terminal, and total stock.",
            "status": gas_status,
        },
        {
            "item": "Import countries",
            "source": "EPDK monthly report Excel appendix",
            "sourceUrl": EPDK_PAGE,
            "pull": "Appendix Table 4 country rows under Spot and Uzun Dönemli İthalat Lisansı.",
            "status": gas_status,
        },
        {
            "item": "Power generation by source",
            "source": "TEİAŞ monthly electricity production-consumption reports",
            "sourceUrl": TEIAS_POWER_PAGE,
            "pull": "Kaynaklara Göre workbook sheet. Hydro = HYDRO. Coal = Hard Coal + Imported Coal plus Lignite. Wind = WIND and Solar = SOLAR when TEİAŞ publishes separate rows. For 2022-2024, TEİAŞ publishes one GEOTHERMAL + WIND + SOLAR row, so individual wind/solar stay blank and the combined row is shown separately. Alternative Fuels = Renew and Wastes / Yenilenebilir + Atık.",
            "status": power_status,
            **({"error": error} if error else {}),
        },
    ]


def empty_payload(status: str, error: str | None = None, power_generation: dict[str, Any] | None = None) -> dict[str, Any]:
    power_generation = power_generation or {
        "status": "source_gap",
        "sourceUrl": TEIAS_POWER_PAGE,
        "rows": [],
    }
    return {
        "schema": "turkey_gas_balance_v1",
        "country": "Turkey",
        "status": status,
        "latest_month": None,
        "cachedAt": datetime.now(timezone.utc).isoformat(),
        "sourcePage": EPDK_PAGE,
        "unit": "bcm/month",
        "rows": [],
        "powerGeneration": power_generation,
        "sourceMap": source_map("source_gap", power_generation.get("status", "source_gap"), error),
        "notes": [
            "No fallback rows are estimated.",
            "Balance equation shown in the app is LNG imports + domestic production + pipeline imports versus EPDK sectorial gas demand.",
        ],
        **({"error": error} if error else {}),
    }


def fetch_bytes(url: str, timeout: int = 60, data: bytes | None = None, headers: dict[str, str] | None = None) -> bytes:
    hdr = {
        "User-Agent": "LNG-TradeOS Turkey gas balance refresh",
        "Accept": "*/*",
    }
    if headers:
        hdr.update(headers)
    req = Request(url, data=data, headers=hdr)
    with urlopen(req, timeout=timeout) as resp:
        return resp.read()


def norm(text: Any) -> str:
    s = html.unescape(str(text or "")).strip().lower()
    s = "".join(ch for ch in unicodedata.normalize("NFKD", s) if not unicodedata.combining(ch))
    table = str.maketrans("çğıöşüâîû", "cgiosuaiu")
    return re.sub(r"\s+", " ", s.translate(table))


def parse_month_label(value: Any, default_year: int | None = None) -> str | None:
    text = norm(value)
    if not text:
        return None
    ym = re.search(r"(20\d{2})[-/\s]+([a-z]+|\d{1,2})", text)
    if ym:
        year = int(ym.group(1))
        raw_m = ym.group(2)
        month = f"{int(raw_m):02d}" if raw_m.isdigit() else MONTHS_TR.get(raw_m)
        return f"{year}-{month}" if month else None
    for name, month in MONTHS_TR.items():
        if name in text and default_year:
            return f"{default_year}-{month}"
    return None


def tr_num(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text or text in {"-", "—", "None"}:
        return None
    text = text.replace("\xa0", "").replace(" ", "")
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    else:
        text = text.replace(",", "")
    try:
        return float(text)
    except ValueError:
        return None


def bcm_from_sm3(value: Any) -> float | None:
    n = tr_num(value)
    return round(n / 1_000_000_000, 6) if n is not None else None


def bcm_from_million_sm3(value: Any) -> float | None:
    n = tr_num(value)
    return round(n / 1000, 6) if n is not None else None


def add(*values: float | None) -> float | None:
    vals = [v for v in values if v is not None]
    return round(sum(vals), 6) if vals else None


def add_required(*values: float | None, digits: int = 6) -> float | None:
    if any(v is None for v in values):
        return None
    return round(sum(v for v in values if v is not None), digits)


def sub(a: float | None, b: float | None) -> float | None:
    if a is None or b is None:
        return None
    return round(a - b, 6)


def report_links(page_html: str) -> list[dict[str, Any]]:
    reports: list[dict[str, Any]] = []
    for li in re.findall(r'<li class="accordion-pop">(.*?)</li>', page_html, re.S | re.I):
        plain = " ".join(re.sub(r"<[^>]+>", " ", html.unescape(li)).split())
        if "Doğal Gaz Piyasası Sektör Raporu" not in plain:
            continue
        m = re.search(r"(Ocak|Şubat|Subat|Mart|Nisan|Mayıs|Mayis|Haziran|Temmuz|Ağustos|Agustos|Eylül|Eylul|Ekim|Kasım|Kasim|Aralık|Aralik)\s+(20\d{2})", plain, re.I)
        if not m:
            continue
        year = int(m.group(2))
        if year < 2025:
            continue
        month = parse_month_label(f"{year} {m.group(1)}")
        if not month:
            continue
        rec: dict[str, Any] = {"month": month, "title": plain, "pdfUrl": None, "excelUrl": None}
        for a in re.finditer(r'<a[^>]+href="([^"]+)"[^>]*title="([^"]*)"[^>]*>(.*?)</a>', li, re.S | re.I):
            href = urljoin(EPDK_PAGE, html.unescape(a.group(1)))
            inner = a.group(3).lower()
            if "excel" in inner:
                rec["excelUrl"] = href
            elif "pdf" in inner:
                rec["pdfUrl"] = href
            elif "word" in inner:
                rec["docxUrl"] = href
        reports.append(rec)
    return sorted(reports, key=lambda x: x["month"])


def cached_gas_rows() -> dict[str, dict[str, Any]]:
    if not OUT.exists():
        return {}
    try:
        payload = json.loads(OUT.read_text(encoding="utf-8"))
    except Exception:
        return {}
    rows = payload.get("rows") or []
    return {str(row.get("month")): row for row in rows if row.get("month")}


def ensure_row(rows: dict[str, dict[str, Any]], month: str) -> dict[str, Any]:
    rows.setdefault(month, {"month": month, "importsByCountry": []})
    return rows[month]


def header_months(ws: Any, row_idx: int, default_year: int) -> dict[int, str]:
    out: dict[int, str] = {}
    for cell in ws[row_idx]:
        mo = parse_month_label(cell.value, default_year)
        if mo:
            out[cell.column] = mo
    return out


def parse_excel(path: Path, rows: dict[str, dict[str, Any]]) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=True)

    # Production: Table 1, Sm3.
    if "Tablo 1-2" in wb.sheetnames:
        ws = wb["Tablo 1-2"]
        year = int(tr_num(ws.cell(3, 2).value) or datetime.now().year)
        months = header_months(ws, 4, year)
        for row in ws.iter_rows(values_only=False):
            if norm(row[0].value) == "genel toplam":
                for col, month in months.items():
                    ensure_row(rows, month)["domesticProductionBcm"] = bcm_from_sm3(row[col - 1].value)

    # Country import rows: Table 4, Sm3.
    if "Tablo 3-4" in wb.sheetnames:
        ws = wb["Tablo 3-4"]
        year = int(tr_num(ws.cell(15, 2).value) or datetime.now().year)
        months = header_months(ws, 16, year)
        group = None
        for row in ws.iter_rows(min_row=17, values_only=False):
            label = str(row[0].value or "").strip()
            label_n = norm(label)
            if not label_n:
                continue
            if label_n in {"spot", "uzun donemli ithalat lisansi"}:
                group = label
                continue
            if label_n == "genel toplam":
                break
            for col, month in months.items():
                vol = bcm_from_sm3(row[col - 1].value)
                if vol is not None:
                    ensure_row(rows, month)["importsByCountry"].append(
                        {"country": label, "licenseGroup": group, "volumeBcm": vol}
                    )

    # Storage stock: Table 6, Sm3.
    if "Tablo 6" in wb.sheetnames:
        ws = wb["Tablo 6"]
        year = int(tr_num(ws.cell(3, 2).value) or datetime.now().year)
        months = header_months(ws, 4, year)
        key_map = {
            "boru gazi": "undergroundStorageStockBcm",
            "lng": "lngTerminalStockBcm",
            "toplam": "totalStorageStockBcm",
        }
        for row in ws.iter_rows(min_row=5, values_only=False):
            key = key_map.get(norm(row[0].value))
            if not key:
                continue
            for col, month in months.items():
                ensure_row(rows, month)[key] = bcm_from_sm3(row[col - 1].value)

    # Sectorial demand: Table 7, million Sm3.
    if "Tablo 7" in wb.sheetnames:
        ws = wb["Tablo 7"]
        months: dict[int, str] = {}
        for cell in ws[4]:
            mo = parse_month_label(cell.value)
            if mo:
                months[cell.column] = mo
        sector_map = {
            "1. donusum/cevrim sektoru": "powerGenerationDemandBcm",
            "4. sanayi sektoru": "industryDemandBcm",
            "6.1. konut": "residentialKonutDemandBcm",
            "genel toplam": "totalDemandBcm",
        }
        for row in ws.iter_rows(min_row=5, values_only=False):
            label = norm(row[0].value)
            key = sector_map.get(label)
            if not key:
                continue
            for col, month in months.items():
                ensure_row(rows, month)[key] = bcm_from_million_sm3(row[col - 1].value)


def parse_pdf_import_split(path: Path) -> dict[str, float | None]:
    try:
        import pdfplumber
    except Exception as exc:  # pragma: no cover - dependency issue in cron
        raise RuntimeError(f"pdfplumber unavailable: {exc}") from exc

    def line_values(label: str, text: str) -> list[float]:
        for line in text.splitlines():
            if norm(line).startswith(norm(label) + " "):
                return [v for v in (tr_num(x) for x in re.findall(r"-?\d[\d.]*,\d+|-?\d+", line)) if v is not None]
        return []

    with pdfplumber.open(str(path)) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            if "Tablo 2.1.2" not in text or "Değişim" not in text:
                continue
            lng_vals = line_values("LNG", text)
            pipe_vals = line_values("Boru Gazı", text)
            # Table layout is previous-year amount/share/current-year amount/share/change.
            return {
                "lngImportsBcm": bcm_from_million_sm3(lng_vals[2] if len(lng_vals) >= 3 else (lng_vals[0] if lng_vals else None)),
                "pipelineImportsBcm": bcm_from_million_sm3(pipe_vals[2] if len(pipe_vals) >= 3 else (pipe_vals[0] if pipe_vals else None)),
            }
    return {"lngImportsBcm": None, "pipelineImportsBcm": None}


W_NS = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}


def docx_cell_text(el: ET.Element) -> str:
    return "".join(t.text or "" for t in el.findall(".//w:t", W_NS)).strip()


def docx_tables(path: Path) -> list[list[list[str]]]:
    with ZipFile(path) as zf:
        root = ET.fromstring(zf.read("word/document.xml"))
    tables: list[list[list[str]]] = []
    for tbl in root.findall(".//w:tbl", W_NS):
        rows: list[list[str]] = []
        for tr in tbl.findall("./w:tr", W_NS):
            rows.append([docx_cell_text(tc) for tc in tr.findall("./w:tc", W_NS)])
        tables.append(rows)
    return tables


def docx_month_from_table(table: list[list[str]]) -> str | None:
    for row in table[:3]:
        for cell in row:
            mo = parse_month_label(cell)
            if mo:
                return mo
    return None


def parse_docx_report(path: Path, rows: dict[str, dict[str, Any]], fallback_month: str) -> None:
    tables = docx_tables(path)

    def current_amount(row: list[str], table: list[list[str]]) -> Any:
        header = norm(" ".join(" ".join(r) for r in table[:2]))
        year = int(fallback_month[:4])
        if str(year - 1) in header and str(year) in header:
            if len(row) >= 6:
                return row[-3]
            if len(row) >= 4:
                return row[3]
        if len(row) >= 6 and "degisim" in header:
            return row[-3]
        return row[1] if len(row) > 1 else None

    def has_labels(table: list[list[str]], *labels: str) -> bool:
        row_labels = {norm(r[0]) for r in table if r}
        return all(norm(label) in row_labels for label in labels)

    for table in tables:
        if not table:
            continue
        flat = " ".join(" ".join(r) for r in table[:8])
        flat_n = norm(flat)
        # DOCX tables often include prior-year comparison columns before the
        # report month. Always anchor parsed values to the report link month.
        month = fallback_month
        row = ensure_row(rows, month)

        # Overview balance table: İthalat, Üretim, Tüketim, Dönem Sonu Stok.
        if any("ithalat" == norm(r[0]) for r in table if r):
            for r in table:
                if len(r) < 3:
                    continue
                label = norm(r[0])
                if label == "uretim":
                    row["domesticProductionBcm"] = bcm_from_million_sm3(r[2])
                elif label == "tuketim":
                    row["totalDemandBcm"] = bcm_from_million_sm3(r[2])
                elif label == "donem sonu stok":
                    row["totalStorageStockBcm"] = bcm_from_million_sm3(r[2])

        # LNG vs pipeline split.
        if "gazin turu" in flat_n and "boru gazi" in flat_n and "lng" in flat_n:
            for r in table:
                label = norm(r[0]) if r else ""
                if label == "lng" and len(r) >= 3:
                    row["lngImportsBcm"] = bcm_from_million_sm3(current_amount(r, table))
                elif label == "boru gazi" and len(r) >= 3:
                    row["pipelineImportsBcm"] = bcm_from_million_sm3(current_amount(r, table))

        # Storage split.
        if "depolama miktari" in flat_n:
            for r in table:
                label = norm(r[0]) if r else ""
                if label == "yeralti depolama":
                    row["undergroundStorageStockBcm"] = bcm_from_million_sm3(current_amount(r, table))
                elif label == "lng terminali":
                    row["lngTerminalStockBcm"] = bcm_from_million_sm3(current_amount(r, table))
                elif label == "genel toplam":
                    row["totalStorageStockBcm"] = bcm_from_million_sm3(current_amount(r, table))

        # Sector table.
        if has_labels(table, "Sanayi Sektörü", "Genel Toplam") and any(norm(r[0]) in {"konut", "6.1. konut"} for r in table if r):
            for r in table:
                label = norm(r[0]) if r else ""
                if len(r) < 3:
                    continue
                current = current_amount(r, table)
                if label in {"donusum/cevrim sektoru", "1. donusum/cevrim sektoru"}:
                    row["powerGenerationDemandBcm"] = bcm_from_million_sm3(current)
                elif label in {"sanayi sektoru", "4. sanayi sektoru"}:
                    row["industryDemandBcm"] = bcm_from_million_sm3(current)
                elif label in {"konut", "6.1. konut"}:
                    row["residentialKonutDemandBcm"] = bcm_from_million_sm3(current)
                elif label == "genel toplam":
                    row["totalDemandBcm"] = bcm_from_million_sm3(current)

        # Import countries by gas type.
        if "ithal edilen ulke" in flat_n and "boru gazi" in flat_n and "lng" in flat_n:
            country_rows: list[dict[str, Any]] = []
            for r in table:
                if not r or len(r) < 5:
                    continue
                country = r[0].strip()
                country_n = norm(country)
                if not country or country_n in {"ithal edilen ulke", "toplam", "genel toplam"}:
                    continue
                # Header has previous-year Boru/LNG then current-year Boru/LNG.
                pipe = bcm_from_million_sm3(r[3])
                lng = bcm_from_million_sm3(r[4])
                if pipe is not None:
                    country_rows.append({"country": country, "licenseGroup": "Boru Gazı", "volumeBcm": pipe})
                if lng is not None:
                    country_rows.append({"country": country, "licenseGroup": "LNG", "volumeBcm": lng})
            if country_rows:
                row["importsByCountry"] = country_rows


def teias_gallery_slug(year: int) -> str:
    return f"{year}-yili-aylik-elektrik-uretim-tuketim-raporlari"


def teias_available_years() -> list[int]:
    try:
        page = fetch_bytes(TEIAS_POWER_PAGE, timeout=30).decode("utf-8", errors="ignore")
        years = sorted({
            int(y)
            for y in re.findall(r"(20\d{2})-yili-aylik-elektrik-uretim-tuketim-raporlari", page)
            if int(y) >= TEIAS_FIXED_HISTORY_START_YEAR
        })
        if years:
            return years
    except Exception:
        pass
    return sorted({*range(TEIAS_FIXED_HISTORY_START_YEAR, TEIAS_FIXED_HISTORY_END_YEAR + 1), datetime.now().year})


def teias_gallery_media(year: int) -> list[dict[str, Any]]:
    raw = fetch_bytes(TEIAS_GALLERY_API + teias_gallery_slug(year), timeout=30).decode("utf-8", errors="ignore")
    data = json.loads(raw)
    if not data.get("success"):
        raise RuntimeError(f"TEİAŞ gallery returned success=false for {year}")
    media = ((data.get("payload") or {}).get("media") or [])
    return [m for m in media if str(m.get("extension", "")).lower() in {"xlsx", "xlsm", "xls"} and m.get("slug")]


def teias_workbook_year(ws: Any, fallback_year: int) -> int:
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 12), values_only=True):
        for cell in row:
            n = tr_num(cell)
            if n and 2000 <= n <= 2100 and float(n).is_integer():
                return int(n)
    return fallback_year


def parse_teias_power_workbook(path: Path, fallback_year: int, source_document: str) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=True)
    if "Kaynaklara Göre" not in wb.sheetnames:
        raise RuntimeError("TEİAŞ workbook missing Kaynaklara Göre sheet")
    ws = wb["Kaynaklara Göre"]
    year = teias_workbook_year(ws, fallback_year)

    month_cols: dict[int, str] = {}
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 15), values_only=False):
        found = 0
        for cell in row:
            label = norm(cell.value)
            month = MONTHS_EN.get(label) or MONTHS_TR.get(label)
            if month:
                month_cols[cell.column] = f"{year}-{month}"
                found += 1
        if found >= 3:
            break
    if not month_cols:
        raise RuntimeError("TEİAŞ workbook month header not found")

    wanted: dict[str, dict[str, Any]] = {
        "hardCoalImportedAsphaltiteGenerationGWh": {
            "label": "Hard Coal + Imported Coal",
            "match": lambda s: "hard coal" in s and "imported coal" in s,
        },
        "ligniteGenerationGWh": {
            "label": "Lignite",
            "match": lambda s: s == "lignite" or s == "linyit",
        },
        "hydroGenerationGWh": {
            "label": "HYDRO",
            "match": lambda s: s == "hydro" or s == "hidrolik",
        },
        "windGenerationGWh": {
            "label": "WIND",
            "match": lambda s: s == "wind" or s == "ruzgar",
        },
        "solarGenerationGWh": {
            "label": "SOLAR",
            "match": lambda s: s == "solar" or s == "gunes",
        },
        "geothermalWindSolarGenerationGWh": {
            "label": "GEOTHERMAL + WIND + SOLAR",
            "match": lambda s: ("geothermal" in s and "wind" in s and "solar" in s) or ("jeotermal" in s and "ruzgar" in s and "gunes" in s),
        },
        "alternativeFuelsGenerationGWh": {
            "label": "Renew and Wastes / Yenilenebilir + Atık",
            "match": lambda s: "renew and wastes" in s or "yenilenebilir + atik" in s,
        },
        "grossGenerationGWh": {
            "label": "GROSS GENERATION",
            "match": lambda s: s == "gross generation" or s == "brut uretim",
        },
    }

    series: dict[str, dict[str, float | None]] = {key: {} for key in wanted}
    labels_found: dict[str, str] = {}
    for row in ws.iter_rows(values_only=False):
        label = norm(row[1].value if len(row) > 1 else "")
        if not label:
            continue
        for key, cfg in wanted.items():
            if cfg["match"](label):
                labels_found[key] = str(row[1].value or cfg["label"]).strip()
                for col, month in month_cols.items():
                    n = tr_num(row[col - 1].value)
                    series[key][month] = round(n, 3) if n is not None else None

    rows: list[dict[str, Any]] = []
    for month in sorted(month_cols.values()):
        hard = series["hardCoalImportedAsphaltiteGenerationGWh"].get(month)
        lignite = series["ligniteGenerationGWh"].get(month)
        rec = {
            "month": month,
            "sourceDocument": source_document,
            "hydroGenerationGWh": series["hydroGenerationGWh"].get(month),
            "hardCoalImportedAsphaltiteGenerationGWh": hard,
            "ligniteGenerationGWh": lignite,
            "coalGenerationGWh": add_required(hard, lignite, digits=3),
            "windGenerationGWh": series["windGenerationGWh"].get(month),
            "solarGenerationGWh": series["solarGenerationGWh"].get(month),
            "geothermalWindSolarGenerationGWh": series["geothermalWindSolarGenerationGWh"].get(month),
            "alternativeFuelsGenerationGWh": series["alternativeFuelsGenerationGWh"].get(month),
            "grossGenerationGWh": series["grossGenerationGWh"].get(month),
            "sourceRows": {key: labels_found.get(key, wanted[key]["label"]) for key in wanted},
        }
        if any(rec.get(k) is not None for k in [
            "hydroGenerationGWh",
            "coalGenerationGWh",
            "windGenerationGWh",
            "solarGenerationGWh",
            "geothermalWindSolarGenerationGWh",
            "alternativeFuelsGenerationGWh",
        ]):
            rows.append(rec)
    return rows


def teias_fixed_months() -> set[str]:
    return {
        f"{year}-{month:02d}"
        for year in range(TEIAS_FIXED_HISTORY_START_YEAR, TEIAS_FIXED_HISTORY_END_YEAR + 1)
        for month in range(1, 13)
    }


def cached_fixed_power_rows() -> dict[str, dict[str, Any]]:
    if not OUT.exists():
        return {}
    try:
        payload = json.loads(OUT.read_text(encoding="utf-8"))
    except Exception:
        return {}
    rows = ((payload.get("powerGeneration") or {}).get("rows") or [])
    fixed: dict[str, dict[str, Any]] = {}
    for row in rows:
        month = str(row.get("month") or "")
        if month in teias_fixed_months():
            fixed[month] = row
    return fixed


def fixed_history_complete(rows: dict[str, dict[str, Any]]) -> bool:
    common_required = {
        "hydroGenerationGWh",
        "hardCoalImportedAsphaltiteGenerationGWh",
        "ligniteGenerationGWh",
        "coalGenerationGWh",
        "alternativeFuelsGenerationGWh",
    }
    expected = teias_fixed_months()
    if not expected.issubset(rows):
        return False
    for month in expected:
        row = rows[month]
        if not all(isinstance(row.get(key), (int, float)) for key in common_required):
            return False
        year = int(month[:4])
        if year <= 2024:
            if not isinstance(row.get("geothermalWindSolarGenerationGWh"), (int, float)):
                return False
        elif not all(isinstance(row.get(key), (int, float)) for key in {"windGenerationGWh", "solarGenerationGWh"}):
            return False
    return True


def teias_years_to_refresh(fixed_complete: bool) -> list[int]:
    years = [year for year in teias_available_years() if year > TEIAS_FIXED_HISTORY_END_YEAR]
    if not fixed_complete:
        years.extend(range(TEIAS_FIXED_HISTORY_START_YEAR, TEIAS_FIXED_HISTORY_END_YEAR + 1))
    return sorted(set(years))


def fetch_power_generation() -> dict[str, Any]:
    payload = {
        "status": "source_gap",
        "sourceUrl": TEIAS_POWER_PAGE,
        "rows": [],
    }
    fixed_rows = cached_fixed_power_rows()
    fixed_complete = fixed_history_complete(fixed_rows)
    by_month: dict[str, dict[str, Any]] = dict(fixed_rows)
    errors: list[str] = []
    try:
        with tempfile.TemporaryDirectory(prefix="lng_turkey_teias_") as tmp:
            tmpdir = Path(tmp)
            for year in teias_years_to_refresh(fixed_complete):
                try:
                    media = teias_gallery_media(year)
                    if not media:
                        errors.append(f"{year}: no TEİAŞ workbook media found")
                        continue
                    for idx, item in enumerate(sorted(media, key=lambda m: str(m.get("created_at", "")))):
                        source_document = item.get("title") or item.get("name") or f"TEİAŞ {year}"
                        path = tmpdir / f"teias_{year}_{idx}.{item.get('extension', 'xlsx')}"
                        path.write_bytes(fetch_bytes(TEIAS_FILE_BASE + item["slug"], timeout=60))
                        for row in parse_teias_power_workbook(path, year, source_document):
                            by_month[row["month"]] = row
                except Exception as exc:
                    errors.append(f"{year}: {exc}")
        rows = [by_month[k] for k in sorted(by_month)]
        if rows:
            fixed_after = {month: by_month[month] for month in teias_fixed_months() if month in by_month}
            fixed_complete_after = fixed_history_complete(fixed_after)
            payload["status"] = "wired" if not errors else "partial"
            payload["latestMonth"] = rows[-1]["month"]
            payload["rows"] = rows
            payload["fixedHistory"] = {
                "startMonth": f"{TEIAS_FIXED_HISTORY_START_YEAR}-01",
                "endMonth": f"{TEIAS_FIXED_HISTORY_END_YEAR}-12",
                "status": "pinned_static" if fixed_complete_after else "incomplete",
                "backfilledThisRun": (not fixed_complete) and fixed_complete_after,
                "rows": len([r for r in rows if str(r.get("month", "")) in teias_fixed_months()]),
            }
            payload["refreshPolicy"] = (
                "TEİAŞ 2022-2025 monthly power-generation rows are fixed history and preserved from the cached JSON. "
                "Weekly refresh only re-scrapes years after 2025 unless the fixed cache is missing or incomplete."
            )
        else:
            payload["status"] = "source_gap"
            payload["error"] = "; ".join(errors) or "No TEİAŞ power generation rows parsed"
        if errors:
            payload["errors"] = errors
    except Exception as exc:  # pragma: no cover - network/source outage
        payload["status"] = "source_gap"
        payload["error"] = str(exc)
    return payload


def build_payload() -> dict[str, Any]:
    power_generation = fetch_power_generation()
    page = fetch_bytes(EPDK_PAGE, timeout=60).decode("utf-8", errors="ignore")
    reports = report_links(page)
    if not reports:
        return empty_payload("source_gap", "No EPDK report links found", power_generation)

    rows: dict[str, dict[str, Any]] = {}
    cached_rows = cached_gas_rows()
    errors: list[str] = []

    with tempfile.TemporaryDirectory(prefix="lng_turkey_") as tmp:
        tmpdir = Path(tmp)
        for report in reports:
            month = report["month"]
            try:
                if report.get("excelUrl"):
                    excel_path = tmpdir / f"epdk_{month}.xlsx"
                    excel_path.write_bytes(fetch_bytes(report["excelUrl"], timeout=120))
                    parse_excel(excel_path, rows)
                if report.get("docxUrl"):
                    docx_path = tmpdir / f"epdk_{month}.docx"
                    docx_path.write_bytes(fetch_bytes(report["docxUrl"], timeout=120))
                    parse_docx_report(docx_path, rows, month)
                if report.get("pdfUrl"):
                    pdf_path = tmpdir / f"epdk_{month}.pdf"
                    pdf_path.write_bytes(fetch_bytes(report["pdfUrl"], timeout=120))
                    ensure_row(rows, month).update(parse_pdf_import_split(pdf_path))
            except Exception as exc:  # pragma: no cover - source-specific failures
                errors.append(f"{month}: {exc}")

    parsed_months = set(rows)
    for month, cached in cached_rows.items():
        if month not in parsed_months and any(err.startswith(f"{month}:") for err in errors):
            cached = dict(cached)
            cached["staleCached"] = True
            rows[month] = cached
            errors.append(f"{month}: preserved previous sourced cache because latest EPDK refresh failed")

    out_rows: list[dict[str, Any]] = []
    for month, row in rows.items():
        lng = row.get("lngImportsBcm")
        pipe = row.get("pipelineImportsBcm")
        domestic = row.get("domesticProductionBcm")
        row["totalImportsBcm"] = add(lng, pipe)
        row["totalSupplyBcm"] = add(lng, domestic, pipe)
        row["trackedDemandBcm"] = add(row.get("powerGenerationDemandBcm"), row.get("industryDemandBcm"), row.get("residentialKonutDemandBcm"))
        row["supplyDemandBalanceBcm"] = sub(row.get("totalSupplyBcm"), row.get("totalDemandBcm"))
        row["importsByCountry"] = sorted(row.get("importsByCountry") or [], key=lambda x: (x.get("licenseGroup") or "", x.get("country") or ""))
        if any(row.get(k) is not None for k in ["totalSupplyBcm", "totalDemandBcm", "lngImportsBcm", "pipelineImportsBcm"]):
            out_rows.append(row)

    out_rows.sort(key=lambda r: r["month"])
    if not out_rows:
        return empty_payload("source_gap", "; ".join(errors) or "No EPDK rows parsed", power_generation)

    gas_status = "wired" if not errors else "partial"
    return {
        "schema": "turkey_gas_balance_v1",
        "country": "Turkey",
        "status": gas_status,
        "latest_month": out_rows[-1]["month"],
        "cachedAt": datetime.now(timezone.utc).isoformat(),
        "sourcePage": EPDK_PAGE,
        "unit": "bcm/month",
        "balanceEquation": "LNG imports + domestic production + pipeline imports = sectorial gas demand",
        "rows": out_rows,
        "powerGeneration": power_generation,
        "sourceMap": source_map(gas_status, power_generation.get("status", "source_gap"), "; ".join(errors) if errors else None),
        "notes": [
            "No fallback rows are estimated.",
            "Supply equation total is LNG imports + domestic production + pipeline imports.",
            "Sectorial gas demand total is the EPDK sector table Genel Toplam.",
            "Power generation/conversion, industry, and residential/Konut are shown as demanded sub-components.",
            "TEİAŞ power-generation source rows are shown separately from the gas balance and are not used to force the EPDK gas equation.",
        ],
        **({"errors": errors} if errors else {}),
    }


def main() -> None:
    OUT.parent.mkdir(parents=True, exist_ok=True)
    try:
        payload = build_payload()
    except Exception as exc:  # pragma: no cover - network/source outage
        payload = empty_payload("source_gap", str(exc))
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    power = payload.get("powerGeneration") or {}
    print(
        f"Turkey gas balance status={payload['status']} rows={len(payload['rows'])} "
        f"latest={payload.get('latest_month')} power={power.get('status')} powerRows={len(power.get('rows') or [])}"
    )


if __name__ == "__main__":
    main()
