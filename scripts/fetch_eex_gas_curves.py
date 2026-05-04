#!/usr/bin/env python3
"""
Refresh data/EEX_Gas_Curves.xlsx from the public EEX Market Data Hub.

The app expects one workbook with hub sheets named THE, PSV, PEG, PVB and ZTP.
Values stay in EUR/MWh in the workbook; the browser converts them to $/MMBtu
with the matching LNG EOD EUR/USD rate.

The scraper only upserts rows that EEX publishes. Existing historical rows are
loaded first and preserved, so a short recent scrape never replaces the history.
"""

from __future__ import annotations

import argparse
import asyncio
import base64
import json
import re
import time
import urllib.parse
import urllib.request
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Dict, Iterable, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
EXCEL_OUT = ROOT / "data" / "EEX_Gas_Curves.xlsx"
EEX_URL = "https://www.eex.com/en/market-data/market-data-hub"
FILTER_URL = "https://api.eex-group.com/pub/customise-widget/filter-data-with-scope"
TABLE_URL = "https://api.eex-group.com/pub/market-data/table-data"
API_HEADERS = {
    "Accept": "application/json, text/javascript, */*; q=0.01",
    "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
    "Origin": "https://www.eex.com",
    "Referer": "https://www.eex.com/",
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 Chrome/124.0.0.0 Safari/537.36"
    ),
}

HUBS = {
    "THE": "German THE Natural Gas Futures",
    "PSV": "Italian PSV Natural Gas Futures",
    "PEG": "French PEG Natural Gas Futures",
    "PVB": "Spanish PVB Natural Gas Futures",
    "ZTP": "Belgian ZTP Natural Gas Futures",
}

MATURITY_LIMITS = {
    "Month": 96,
    "Quarter": 32,
    "Season": 18,
    "Year": 10,
}
MATURITY_ORDER = {"Month": 0, "Quarter": 1, "Season": 2, "Year": 3}

C_NAVY = "1F3864"
C_BLUE = "2E75B6"
C_WHITE = "FFFFFF"
C_GREY = "F2F2F2"

BookData = Dict[str, Dict[str, Dict[str, float]]]


def parse_float(value: object) -> Optional[float]:
    if value is None:
        return None
    clean = re.sub(r"[^\d.,-]", "", str(value))
    if not clean:
        return None
    if clean.count(",") == 1 and "." not in clean:
        clean = clean.replace(",", ".")
    elif "," in clean and "." in clean:
        clean = (
            clean.replace(".", "").replace(",", ".")
            if clean.rfind(",") > clean.rfind(".")
            else clean.replace(",", "")
        )
    try:
        return float(clean)
    except ValueError:
        return None


def parse_trade_date(value: object) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", text)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    for fmt in ("%d %b %Y", "%d %B %Y"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return None


def normalise_tenor(raw: object) -> str:
    text = re.sub(r"\s+", " ", str(raw or "").strip())
    months = {
        "January": "Jan",
        "February": "Feb",
        "March": "Mar",
        "April": "Apr",
        "May": "May",
        "June": "Jun",
        "July": "Jul",
        "August": "Aug",
        "September": "Sep",
        "October": "Oct",
        "November": "Nov",
        "December": "Dec",
        "Sept": "Sep",
    }
    abbr = {v: v for v in months.values()}

    m = re.fullmatch(r"([A-Za-z]+)[ -](\d{2}|\d{4})", text)
    if m:
        name = m.group(1)
        year = m.group(2)
        yy = year[-2:]
        if name in months or name in abbr:
            return f"{months.get(name, abbr.get(name, name[:3]))}-{yy}"
        if name.lower().startswith("win"):
            return f"Win-{yy}"
        if name.lower().startswith("sum"):
            return f"Sum-{yy}"
        if name.lower().startswith("cal"):
            return f"Cal-{yy}"

    m = re.fullmatch(r"Q([1-4])[ -]?(\d{2}|\d{4})", text, re.I)
    if m:
        return f"Q{m.group(1)}-{m.group(2)[-2:]}"

    m = re.fullmatch(r"(Winter|Summer|Season|Year|Calendar Year)\s+(\d{4})", text, re.I)
    if m:
        prefix = m.group(1).lower()
        yy = m.group(2)[-2:]
        if prefix.startswith("winter"):
            return f"Win-{yy}"
        if prefix.startswith("summer"):
            return f"Sum-{yy}"
        return f"Cal-{yy}"

    m = re.fullmatch(r"(\d{4})", text)
    if m:
        return f"Cal-{m.group(1)[-2:]}"

    return text


def tenor_from_filter_row(record: dict) -> str:
    year = int(record["displayYear"])
    yy = str(year)[-2:]
    maturity_type = record["maturityType"]
    month = record.get("displayMonth")
    if maturity_type == "Month":
        month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        return f"{month_names[int(month) - 1]}-{yy}"
    if maturity_type == "Quarter":
        return f"Q{int(record['displayQuarter'])}-{yy}"
    if maturity_type == "Season":
        return f"{'Sum' if int(month) == 4 else 'Win'}-{yy}"
    if maturity_type == "Year":
        return f"Cal-{yy}"
    return normalise_tenor(record["maturity"])


def tenor_sort_key(tenor: str) -> tuple:
    months = {
        "Jan": 1,
        "Feb": 2,
        "Mar": 3,
        "Apr": 4,
        "May": 5,
        "Jun": 6,
        "Jul": 7,
        "Aug": 8,
        "Sep": 9,
        "Oct": 10,
        "Nov": 11,
        "Dec": 12,
    }
    m = re.fullmatch(r"([A-Z][a-z]{2})-(\d{2})", tenor)
    if m:
        return (2000 + int(m.group(2)), months.get(m.group(1), 0), 0)
    m = re.fullmatch(r"Q([1-4])-(\d{2})", tenor)
    if m:
        return (2000 + int(m.group(2)), 20 + int(m.group(1)), 0)
    m = re.fullmatch(r"(Sum|Win)-(\d{2})", tenor)
    if m:
        return (2000 + int(m.group(2)), 30 if m.group(1) == "Sum" else 31, 0)
    m = re.fullmatch(r"Cal-(\d{2})", tenor)
    if m:
        return (2000 + int(m.group(1)), 40, 0)
    return (9999, 99, tenor)


def load_existing(path: Path = EXCEL_OUT) -> BookData:
    data: BookData = {hub: {} for hub in HUBS}
    if not path.exists():
        return data
    wb = load_workbook(path, data_only=True, read_only=True)
    for hub in HUBS:
        if hub not in wb.sheetnames:
            continue
        ws = wb[hub]
        headers = [normalise_tenor(c.value) if c.value else "" for c in ws[2]]
        for row in ws.iter_rows(min_row=3, values_only=True):
            ds = parse_trade_date(row[0])
            if not ds:
                continue
            bucket = data[hub].setdefault(ds, {})
            for idx, value in enumerate(row[1:], start=1):
                tenor = headers[idx] if idx < len(headers) else ""
                val = parse_float(value)
                if tenor and val is not None:
                    bucket[tenor] = val
    return data


def api_json(url: str, data: Optional[dict] = None) -> dict:
    encoded = None
    if data is not None:
        encoded = urllib.parse.urlencode(data).encode()
    req = urllib.request.Request(url, data=encoded, headers=API_HEADERS, method="POST" if data is not None else "GET")
    with urllib.request.urlopen(req, timeout=30) as resp:
        body = resp.read().decode("utf-8")
    return json.loads(body)


def load_filter_rows() -> List[dict]:
    contracts = [
        {
            "commodity": "All",
            "pricing": "All",
            "area": "All",
            "product": "All",
            "productSpecific": "All",
            "maturityType": "All",
        }
    ]
    token = base64.b64encode(json.dumps(contracts, separators=(",", ":")).encode()).decode()
    payload = api_json(f"{FILTER_URL}?data={urllib.parse.quote(token)}", {"data": token})
    header = payload["header"]
    out = []
    for row in payload.get("data", []):
        out.append({key: row[idx] if idx < len(row) else None for idx, key in enumerate(header)})
    return out


def eex_contracts_for_hubs(hubs: Iterable[str], tenor_limit: Optional[int] = None) -> List[dict]:
    wanted = set(hubs)
    records = []
    for row in load_filter_rows():
        if row.get("commodity") != "NATGAS" or row.get("pricing") != "F":
            continue
        if row.get("area") not in wanted or row.get("product") != "Physical":
            continue
        if row.get("maturityType") not in MATURITY_ORDER:
            continue
        row["tenor"] = tenor_from_filter_row(row)
        records.append(row)

    records.sort(
        key=lambda r: (
            r["area"],
            MATURITY_ORDER[r["maturityType"]],
            tenor_sort_key(r["tenor"]),
        )
    )
    if tenor_limit:
        limited = []
        counts = {}
        for row in records:
            key = (row["area"], row["maturityType"])
            counts[key] = counts.get(key, 0) + 1
            if counts[key] <= tenor_limit:
                limited.append(row)
        records = limited
    return records


def fetch_table_rows(record: dict, start: str, end: str) -> List[dict]:
    params = {
        "shortCode": record["shortCode"],
        "commodity": "NATGAS",
        "pricing": "F",
        "area": record["area"],
        "product": "Physical",
        "maturity": record["maturity"],
        "startDate": start,
        "endDate": end,
        "maturityType": record["maturityType"],
        "isRolling": "true",
    }
    payload = api_json(TABLE_URL + "?" + urllib.parse.urlencode(params))
    header = payload.get("header", [])
    idx = {name: header.index(name) for name in header}
    out = []
    for row in payload.get("data", []):
        trade_date = row[idx["tradeDate"]] if "tradeDate" in idx else None
        settlement = row[idx["settlPx"]] if "settlPx" in idx else None
        if trade_date and settlement is not None:
            out.append(
                {
                    "trade_date": str(trade_date)[:10],
                    "hub": record["area"],
                    "tenor": record["tenor"],
                    "settlement": settlement,
                }
            )
    return out


def scrape_api(hubs: Iterable[str], tenor_limit: Optional[int] = None) -> List[dict]:
    start = (date.today() - timedelta(days=45)).isoformat()
    end = date.today().isoformat()
    contracts = eex_contracts_for_hubs(hubs, tenor_limit=tenor_limit)
    rows = []
    print(f"[EEX] API contracts: {len(contracts)} · range {start} to {end}")
    for pos, record in enumerate(contracts, start=1):
        try:
            got = fetch_table_rows(record, start, end)
            rows.extend(got)
            if pos % 25 == 0 or got:
                latest = max((r["trade_date"] for r in got), default="gap")
                print(f"[EEX] {record['area']} {record['tenor']}: {len(got)} rows · latest {latest}")
            time.sleep(0.03)
        except Exception as exc:
            print(f"[EEX] {record['area']} {record['tenor']}: source gap ({exc})")
    return rows


async def select_by_label_or_text(select, wanted: str) -> bool:
    try:
        await select.select_option(label=wanted)
        return True
    except Exception:
        pass
    opts = [o.strip() for o in await select.locator("option").all_inner_texts()]
    for opt in opts:
        if opt == wanted or wanted.lower() in opt.lower():
            await select.select_option(label=opt)
            return True
    return False


async def dismiss_popups(page) -> None:
    await page.evaluate(
        """
        (() => {
          document.querySelectorAll('#popup,.eex-popup,.modal.show,.modal-backdrop,.cc-window')
            .forEach(e => e.remove());
          document.body.classList.remove('modal-open');
          document.body.style.overflow = '';
        })()
        """
    )


async def click_go(page) -> None:
    from playwright.async_api import TimeoutError as PlaywrightTimeout

    for selector in ("button:has-text('GO')", "button.btn-go", "#btn-go", "input[value='GO']"):
        try:
            button = page.locator(selector).first
            if await button.is_visible(timeout=800):
                await button.click()
                try:
                    await page.wait_for_load_state("networkidle", timeout=8_000)
                except PlaywrightTimeout:
                    pass
                await asyncio.sleep(0.8)
                return
        except Exception:
            pass
    await asyncio.sleep(0.8)


async def read_settlement_table(page, hub: str, tenor: str) -> List[dict]:
    rows: List[dict] = []
    try:
        table = page.locator("table.table-hover").first
        tr_rows = table.locator("tbody tr")
        row_count = await tr_rows.count()
        if row_count == 0:
            tr_rows = table.locator("tr:not(:first-child)")
            row_count = await tr_rows.count()

        for idx in range(row_count):
            cells = tr_rows.nth(idx).locator("td")
            if await cells.count() < 4:
                continue
            trade_date = parse_trade_date((await cells.nth(0).inner_text()).strip())
            settlement = parse_float((await cells.nth(3).inner_text()).strip())
            if not trade_date or settlement is None:
                continue
            rows.append(
                {
                    "trade_date": trade_date,
                    "hub": hub,
                    "tenor": tenor,
                    "settlement": settlement,
                }
            )
    except Exception:
        return rows
    return rows


async def scrape_hub(page, hub: str, tenor_limit: Optional[int] = None) -> List[dict]:
    from playwright.async_api import TimeoutError as PlaywrightTimeout

    print(f"[EEX] {hub}: loading")
    await page.goto(EEX_URL, wait_until="networkidle", timeout=45_000)
    await asyncio.sleep(1.2)
    await dismiss_popups(page)
    selects = page.locator("select.form-control")
    if await selects.count() < 10:
        raise RuntimeError(f"{hub}: EEX dropdown structure not available")

    await select_by_label_or_text(selects.nth(0), "Natural Gas")
    await asyncio.sleep(1.0)
    await select_by_label_or_text(selects.nth(1), "Futures")
    await asyncio.sleep(1.0)
    if not await select_by_label_or_text(selects.nth(2), hub):
        raise RuntimeError(f"{hub}: hub option not found")
    await asyncio.sleep(1.0)

    product_options = [o.strip() for o in await selects.nth(3).locator("option").all_inner_texts()]
    product = "Physical" if "Physical" in product_options else "Financial"
    await select_by_label_or_text(selects.nth(3), product)
    await asyncio.sleep(1.0)

    out: List[dict] = []
    for maturity, default_limit in MATURITY_LIMITS.items():
        if not await select_by_label_or_text(selects.nth(8), maturity):
            print(f"[EEX] {hub}: maturity {maturity} unavailable")
            continue
        await asyncio.sleep(1.0)

        raw_options = [o.strip() for o in await selects.nth(9).locator("option").all_inner_texts()]
        raw_options = [o for o in raw_options if o and "select" not in o.lower()]
        limit = tenor_limit or default_limit
        raw_options = raw_options[:limit]
        print(f"[EEX] {hub}: {maturity} {len(raw_options)} tenors")

        for pos, raw_tenor in enumerate(raw_options, start=1):
            tenor = normalise_tenor(raw_tenor)
            try:
                await selects.nth(9).select_option(label=raw_tenor)
            except Exception:
                continue
            try:
                await page.wait_for_load_state("networkidle", timeout=6_000)
            except PlaywrightTimeout:
                pass
            await click_go(page)
            rows = await read_settlement_table(page, hub, tenor)
            out.extend(rows)
            if pos % 12 == 0:
                print(f"[EEX] {hub}: {maturity} {pos}/{len(raw_options)}")

    dates = sorted({r["trade_date"] for r in out})
    print(f"[EEX] {hub}: {len(out)} rows, latest {dates[-1] if dates else 'none'}")
    return out


async def scrape(hubs: Iterable[str], tenor_limit: Optional[int] = None) -> List[dict]:
    from playwright.async_api import async_playwright

    rows: List[dict] = []
    async with async_playwright() as pw:
        browser = await pw.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-dev-shm-usage"],
        )
        context = await browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 Chrome/124.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1440, "height": 900},
            locale="en-GB",
        )
        page = await context.new_page()
        for hub in hubs:
            try:
                rows.extend(await scrape_hub(page, hub, tenor_limit=tenor_limit))
            except Exception as exc:
                print(f"[EEX] {hub}: failed: {exc}")
        await browser.close()
    return rows


def merge_rows(data: BookData, rows: Iterable[dict]) -> int:
    count = 0
    for row in rows:
        hub = row["hub"]
        if hub not in HUBS:
            continue
        ds = row["trade_date"]
        tenor = normalise_tenor(row["tenor"])
        value = parse_float(row["settlement"])
        if not ds or not tenor or value is None:
            continue
        data.setdefault(hub, {}).setdefault(ds, {})[tenor] = round(value, 3)
        count += 1
    return count


def style_header(cell, bg=C_NAVY, fg=C_WHITE) -> None:
    cell.font = Font(name="Calibri", bold=True, color=fg, size=10)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def export_excel(data: BookData, path: Path = EXCEL_OUT) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    del wb["Sheet"]

    ws = wb.create_sheet("SUMMARY")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:F1")
    title = ws["A1"]
    title.value = "EEX EUROPEAN GAS FUTURES  ·  Settlement Curve History"
    title.font = Font(name="Calibri", bold=True, size=16, color=C_WHITE)
    title.fill = PatternFill("solid", start_color=C_NAVY)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A2:F2")
    sub = ws["A2"]
    sub.value = (
        f"Generated: {datetime.now(timezone.utc).strftime('%d %b %Y %H:%M')} UTC"
        "  ·  Source: EEX Market Data Hub (public)  ·  EUR/MWh"
    )
    sub.font = Font(name="Calibri", size=9, italic=True, color=C_WHITE)
    sub.fill = PatternFill("solid", start_color=C_BLUE)
    sub.alignment = Alignment(horizontal="center")

    for ci, header in enumerate(["Hub", "Full Name", "Days", "From", "To", "Rows"], start=1):
        style_header(ws.cell(row=4, column=ci, value=header), bg=C_BLUE)

    for ri, hub in enumerate(HUBS, start=5):
        dates = sorted(data.get(hub, {}))
        rows = sum(len(data.get(hub, {}).get(ds, {})) for ds in dates)
        values = [
            hub,
            HUBS[hub],
            len(dates),
            dates[0] if dates else "—",
            dates[-1] if dates else "—",
            rows,
        ]
        bg = C_GREY if ri % 2 == 0 else C_WHITE
        for ci, value in enumerate(values, start=1):
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.font = Font(name="Calibri", size=10)
            cell.fill = PatternFill("solid", start_color=bg)
            cell.alignment = Alignment(horizontal="center" if ci != 2 else "left")
    for ci, width in enumerate([8, 36, 8, 14, 14, 8], start=1):
        ws.column_dimensions[get_column_letter(ci)].width = width

    for hub in HUBS:
        sheet = wb.create_sheet(hub)
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "B3"
        sheet.merge_cells("A1:E1")
        heading = sheet["A1"]
        heading.value = f"{hub}  ·  {HUBS[hub]}  ·  EUR/MWh"
        heading.font = Font(name="Calibri", bold=True, size=13, color=C_WHITE)
        heading.fill = PatternFill("solid", start_color=C_NAVY)
        heading.alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[1].height = 26

        dates = sorted(data.get(hub, {}), reverse=True)
        tenors = sorted({t for ds in data.get(hub, {}).values() for t in ds}, key=tenor_sort_key)
        style_header(sheet.cell(row=2, column=1, value="Trade Date"), bg=C_BLUE)
        sheet.column_dimensions["A"].width = 16
        for ci, tenor in enumerate(tenors, start=2):
            style_header(sheet.cell(row=2, column=ci, value=tenor), bg=C_BLUE)
            sheet.column_dimensions[get_column_letter(ci)].width = 10

        for ri, ds in enumerate(dates, start=3):
            bg = C_WHITE if ri % 2 else C_GREY
            dt = datetime.strptime(ds, "%Y-%m-%d")
            date_cell = sheet.cell(row=ri, column=1, value=dt.strftime("%d %b %Y"))
            date_cell.font = Font(name="Calibri", size=10, bold=True)
            date_cell.fill = PatternFill("solid", start_color=bg)
            for ci, tenor in enumerate(tenors, start=2):
                value = data[hub][ds].get(tenor)
                cell = sheet.cell(row=ri, column=ci, value=value)
                cell.font = Font(name="Calibri", size=10)
                cell.fill = PatternFill("solid", start_color=bg)
                cell.number_format = "#,##0.000"
                cell.alignment = Alignment(horizontal="right")

    wb.save(path)
    print(f"[EEX] saved {path}")


def print_status(data: BookData) -> None:
    print(f"{'Hub':<6} {'Days':>5} {'From':<12} {'To':<12} {'Rows':>7}")
    print("-" * 48)
    for hub in HUBS:
        dates = sorted(data.get(hub, {}))
        rows = sum(len(data.get(hub, {}).get(ds, {})) for ds in dates)
        print(
            f"{hub:<6} {len(dates):>5} "
            f"{(dates[0] if dates else '—'):<12} "
            f"{(dates[-1] if dates else '—'):<12} {rows:>7}"
        )


def main() -> None:
    parser = argparse.ArgumentParser(description="Refresh EEX gas curve workbook")
    parser.add_argument("--hubs", nargs="+", default=list(HUBS), help="Hub sheet names to scrape")
    parser.add_argument("--status", action="store_true", help="Show workbook status and exit")
    parser.add_argument("--export-only", action="store_true", help="Re-export existing workbook without scraping")
    parser.add_argument("--tenor-limit", type=int, default=None, help="Debug limit per maturity bucket")
    args = parser.parse_args()

    data = load_existing()
    if args.status:
        print_status(data)
        return
    if not args.export_only:
        hubs = [h.upper() for h in args.hubs if h.upper() in HUBS]
        scraped = scrape_api(hubs, tenor_limit=args.tenor_limit)
        upserts = merge_rows(data, scraped)
        print(f"[EEX] scraped rows merged: {upserts}")
        if not upserts:
            raise SystemExit("No EEX rows scraped; keeping existing workbook unchanged.")
    export_excel(data)
    print_status(data)


if __name__ == "__main__":
    main()
