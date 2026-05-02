#!/usr/bin/env python3
"""
Korea daily data scraper for LNG TradeOS.

Sources (all globally accessible, no API key, no Korean-IP requirement):
  - KHNP open NPP operations portal — LIVE per-reactor output % and MWe
    (npp.khnp.co.kr is reachable globally; only the cms. subdomain on the
    realTimeMgr path is blocked. The npp. subdomain serves the same data
    via a JSON API used by their public dashboard, refreshed every ~3 min.)
  - ECB euro reference rates — KRW/USD via cross-rate
  - KOGAS official power-sector tariff page — current monthly tariff

KOGAS imports remain manual until either a Korean proxy lands or the user
enters values. DART quarterly filings need the user's free OpenDART API key
(1-week approval).

Output: data/korea.json
Cadence: daily 17:30 UTC (configurable; KHNP refreshes every ~3 min, but
daily commits are fine — site-level counts only change on outages, which
are rare. Could move to hourly later if needed.)
"""
import calendar
import json
import os
import re
import sys
import time
import xml.etree.ElementTree as ET
from datetime import datetime, timezone, timedelta


# ────────────────────────────────────────────────────────────────────
# Korean nuclear fleet nameplate capacity timeline (IAEA PRIS-anchored).
# Used to convert Ember monthly TWh → monthly utilization (capacity factor).
#
# Events (commercial-operation / permanent-shutdown dates):
#   2019-08-29  Shin Kori 4   commercial (+1.416 GW)
#   2019-12-24  Wolsong 1     retired    (-0.679 GW)
#   2022-12-07  Shin Hanul 1  commercial (+1.418 GW)
#   2023-04-08  Kori 2        retired    (-0.650 GW)
#   2024-04-05  Shin Hanul 2  commercial (+1.418 GW)
#   2025-03-28  Shin Kori 5   commercial (+1.416 GW)
# ────────────────────────────────────────────────────────────────────
def korea_nuc_capacity_gw(date_str):
    """Operating Korean nuclear nameplate (GW) for a YYYY-MM-DD month."""
    ym = int(date_str[:4]) * 100 + int(date_str[5:7])
    if ym < 201909: return 22.67   # pre-Shin-Kori-4
    if ym < 201912: return 24.08   # SK4 in, Wolsong 1 still
    if ym < 202212: return 23.40   # Wolsong 1 retired
    if ym < 202304: return 24.82   # Shin Hanul 1 added
    if ym < 202404: return 24.17   # Kori 2 retired
    if ym < 202503: return 25.59   # Shin Hanul 2 added
    return 27.00                    # Shin Kori 5 added

# Prefer requests (more robust SSL session handling than urllib for sites
# like IAEA PRIS that flake on TLS handshakes). Fall back to urllib if not.
try:
    import requests as _requests
    _USE_REQUESTS = True
except ImportError:
    import urllib.request
    _USE_REQUESTS = False

OUT_PATH = os.path.join(os.path.dirname(__file__), '..', 'data', 'korea.json')
UA = 'Mozilla/5.0 (LNG-TradeOS Korea scraper)'

# Map PRIS reactor name prefix → site label shown in the dashboard.
# SHIN- variants merged into their parent site to mirror Korean industry usage.
SITE_MAP = {
    'HANUL':       'Hanul',
    'SHIN-HANUL':  'Hanul',
    'HANBIT':      'Hanbit',
    'KORI':        'Kori',
    'SHIN-KORI':   'Kori',
    'SAEUL':       'Saeul',
    'WOLSONG':     'Wolsong',
    'SHIN-WOLSONG':'Wolsong',
}


def fetch(url, timeout=20, retries=4):
    """HTTP GET with retries.

    Python's default SSL stack rejects the cipher suite IAEA PRIS serves
    (SSLZeroReturnError on every attempt), but the system curl works
    cleanly against the same URL. We try Python first (fast, in-process)
    and fall back to a curl subprocess if it fails — this keeps the
    common case efficient while making PRIS reliably reachable.
    """
    import subprocess
    last = None
    for attempt in range(retries):
        try:
            if _USE_REQUESTS:
                r = _requests.get(url, headers={'User-Agent': UA}, timeout=timeout)
                r.raise_for_status()
                return r.text
            req = urllib.request.Request(url, headers={'User-Agent': UA})
            with urllib.request.urlopen(req, timeout=timeout) as r:
                return r.read().decode('utf-8')
        except Exception as e:
            last = e
            if attempt < retries - 1:
                time.sleep(2 ** attempt)
    # All Python attempts failed — try curl subprocess as a final fallback.
    try:
        out = subprocess.run(
            ['curl', '-fsS', '-A', UA, '--max-time', str(timeout), url],
            check=True, capture_output=True, text=True, timeout=timeout + 5,
        )
        return out.stdout
    except Exception as e:
        raise last if last else e


# KHNP unit codes per site. unitCd is the per-reactor identifier in the
# /branch-operation-info-by-plant API. Discovered by scraping each site's
# unit dropdown on npp.khnp.co.kr/ON004004002002NNN. branchCd3 (ho) is
# unused by the API for unit filtering — kept here as a label only.
KHNP_SITES = [
    {'site': 'Kori',    'branchCd': 'BR0302', 'units': [
        ('2112', 'Kori-2', 2),       ('2123', 'Kori-3', 3),
        ('2124', 'Kori-4', 4),       ('2135', 'Shin-Kori-1', 1),
        ('2136', 'Shin-Kori-2', 2),
    ]},
    {'site': 'Hanbit',  'branchCd': 'BR0303', 'units': [
        ('2311', 'Hanbit-1', 1),     ('2312', 'Hanbit-2', 2),
        ('2323', 'Hanbit-3', 3),     ('2324', 'Hanbit-4', 4),
        ('2335', 'Hanbit-5', 5),     ('2336', 'Hanbit-6', 6),
    ]},
    {'site': 'Hanul',   'branchCd': 'BR0304', 'units': [
        ('2411', 'Hanul-1', 1),      ('2412', 'Hanul-2', 2),
        ('2423', 'Hanul-3', 3),      ('2424', 'Hanul-4', 4),
        ('2435', 'Hanul-5', 5),      ('2436', 'Hanul-6', 6),
        ('2711', 'Shin-Hanul-1', 1), ('2712', 'Shin-Hanul-2', 2),
    ]},
    {'site': 'Wolsong', 'branchCd': 'BR0305', 'units': [
        ('2212', 'Wolsong-2', 2),    ('2223', 'Wolsong-3', 3),
        ('2224', 'Wolsong-4', 4),    ('2235', 'Shin-Wolsong-1', 1),
        ('2236', 'Shin-Wolsong-2', 2),
    ]},
    {'site': 'Saeul',   'branchCd': 'BR0312', 'units': [
        ('2811', 'Saeul-1', 1),      ('2812', 'Saeul-2', 2),
    ]},
]


def fetch_khnp_unit(branch_cd, unit_cd, ho, session_cookies=None):
    """Hit KHNP's /branch-operation-info-by-plant for a single reactor.

    Returns None on failure or empty response (e.g. unit shutdown / not in
    KHNP's monitored list any more). Live data refreshes every ~3 minutes
    on KHNP's side; we capture whatever the latest snapshot is.
    """
    import json as _json
    url = 'https://npp.khnp.co.kr/branch-operation-info-by-plant'
    payload = _json.dumps({'branchCd': branch_cd, 'branchCd2': unit_cd, 'branchCd3': str(ho)})
    headers = {
        'User-Agent': UA,
        'Accept': 'application/json, */*; q=0.01',
        'Content-Type': 'application/json',
        'X-Requested-With': 'XMLHttpRequest',
        'Referer': 'https://npp.khnp.co.kr/ON004004002002001',
        'Origin': 'https://npp.khnp.co.kr',
    }
    try:
        if _USE_REQUESTS:
            r = _requests.post(url, data=payload, headers=headers,
                               cookies=session_cookies, timeout=15)
            if r.status_code != 200 or not r.text.strip():
                return None
            return r.json()
        # urllib fallback
        req = urllib.request.Request(url, data=payload.encode('utf-8'), headers=headers, method='POST')
        with urllib.request.urlopen(req, timeout=15) as resp:
            body = resp.read().decode('utf-8')
        return _json.loads(body) if body.strip() else None
    except Exception:
        return None


def fetch_khnp_session():
    """Hit the npp.khnp.co.kr root once to obtain the WMONID session cookie
    that the API requires. Without this, the API returns empty 200s."""
    if not _USE_REQUESTS:
        return None
    s = _requests.Session()
    s.headers.update({'User-Agent': UA})
    try:
        s.get('https://npp.khnp.co.kr/', timeout=15)
        return s
    except Exception:
        return None


def fetch_khnp_live():
    """Pull live per-reactor output from KHNP's open ops portal.

    Returns a structure compatible with the front-end (reactors[], bySite[],
    totalOnlineGW, onlineCount, totalCount), plus the underlying KHNP
    timestamps so the front-end can display data freshness honestly.
    """
    session = fetch_khnp_session()
    cookies = session.cookies if session else None

    reactors = []
    timestamps = []
    for site in KHNP_SITES:
        for unit_cd, name, ho in site['units']:
            data = fetch_khnp_unit(site['branchCd'], unit_cd, ho,
                                    session_cookies=cookies)
            if not data:
                # Append a stub with status unknown so site totals stay sensible
                reactors.append({
                    'name': name, 'site': site['site'], 'unit_cd': unit_cd,
                    'status': 'Unknown', 'output_pct': None, 'output_mwe': None,
                    'asOf': None,
                })
                continue
            udo = data.get('unitDetailOutput') or {}
            no1 = udo.get('NO_1') or {}
            no8 = udo.get('NO_8') or {}
            ul = (data.get('unitInfoList') or [{}])[0]
            # Validate the response is for the requested reactor. plantCd
            # in unitDetailOutput is the authoritative identifier — the
            # unitInfoList[0] sometimes returns the first unit at the
            # site instead of the requested one (KHNP server-side quirk).
            api_unit = udo.get('plantCd')
            if api_unit and api_unit != unit_cd:
                reactors.append({
                    'name': name, 'site': site['site'], 'unit_cd': unit_cd,
                    'status': 'Unknown', 'output_pct': None, 'output_mwe': None,
                    'asOf': None,
                })
                continue
            try:
                output_pct = float(no1.get('VALUE')) if no1.get('VALUE') is not None else None
            except (TypeError, ValueError):
                output_pct = None
            try:
                output_mwe = float(no8.get('VALUE')) if no8.get('VALUE') is not None else None
            except (TypeError, ValueError):
                output_mwe = None
            # Status: derive from generator output, not the kh_status field
            # (which is unreliable for some units). Unit clearly operational
            # if MWe > 50, clearly offline if MWe is null/0.
            kh_status = ul.get('status') or ''
            if output_mwe is not None and output_mwe > 50:
                status = 'Operational'
            elif output_mwe is None or output_mwe < 5:
                status = 'Maintenance' if kh_status == 'KH1202' else 'Offline'
            else:
                status = 'Partial'
            ts = no1.get('TIME') or no8.get('TIME')
            if ts:
                timestamps.append(ts)
            reactors.append({
                'name':       name,
                'site':       site['site'],
                'unit_cd':    unit_cd,
                'status':     status,
                'output_pct': output_pct,
                'output_mwe': output_mwe,
                'asOf':       ts,
            })

    # Aggregate per site
    sites_dict = {}
    for r in reactors:
        s = sites_dict.setdefault(r['site'], {
            'site': r['site'], 'reactors': [], 'totalCount': 0, 'online': 0,
            'totalCap_GW': 0.0, 'onlineCap_GW': 0.0,
        })
        s['reactors'].append(r['name'])
        s['totalCount'] += 1
        # Totalcap = sum of all reactor MWe (when reported, these are nameplate-ish).
        # Onlinecap = sum of MWe for operational reactors.
        if r['output_mwe'] and r['output_mwe'] > 0:
            s['totalCap_GW'] += r['output_mwe'] / 1000.0
        if r['status'] == 'Operational':
            s['online'] += 1
            if r['output_mwe']:
                s['onlineCap_GW'] += r['output_mwe'] / 1000.0

    by_site = sorted(sites_dict.values(), key=lambda x: -x['onlineCap_GW'])
    for s in by_site:
        s['totalCap_GW'] = round(s['totalCap_GW'], 2)
        s['onlineCap_GW'] = round(s['onlineCap_GW'], 2)

    online_total_gw = round(sum(s['onlineCap_GW'] for s in by_site), 2)
    online_count = sum(s['online'] for s in by_site)
    total_count = sum(s['totalCount'] for s in by_site)
    # Latest underlying KHNP timestamp — what the user sees as "data as of"
    latest_ts = max(timestamps) if timestamps else None

    return {
        'asOf':         latest_ts,  # KHNP's own timestamp, NOT our scrape time
        'scrapedAt':    datetime.now(timezone.utc).isoformat(timespec='seconds'),
        'source':       'KHNP Open NPP Operations Info',
        'sourceUrl':    'https://npp.khnp.co.kr/',
        'reactors':     reactors,
        'bySite':       by_site,
        'totalOnlineGW':online_total_gw,
        'onlineCount':  online_count,
        'totalCount':   total_count,
    }


def fetch_khnp_annual_util():
    """Annual fleet utilization + capacity factor from KHNP.

    Source: https://npp.khnp.co.kr/ON004004002001003 — '이용률·가동률' page.
    Single server-rendered table with a year-by-year history (2007 onwards).
    Returns: { asOf, byYear: [{year, util, capFactor}, ...], source }
    """
    url = 'https://npp.khnp.co.kr/ON004004002001003'
    html = fetch(url)
    tables = re.findall(r'<table[^>]*>.*?</table>', html, re.DOTALL)
    if not tables:
        return None
    tbl = tables[0]
    # Header: 구분 + year columns
    th_all = [re.sub(r'<[^>]+>', '', t).strip()
              for t in re.findall(r'<th[^>]*>(.*?)</th>', tbl, re.DOTALL)]
    years = [int(t) for t in th_all if t.isdigit() and len(t) == 4]
    rows = re.findall(r'<tr[^>]*>(.*?)</tr>', tbl, re.DOTALL)
    util_row, cap_row = None, None
    for r in rows:
        cells = [re.sub(r'<[^>]+>', '', c).strip()
                 for c in re.findall(r'<t[dh][^>]*>(.*?)</t[dh]>', r, re.DOTALL)]
        if not cells:
            continue
        if '이용률' in cells[0]:
            util_row = cells[1:]
        elif '가동률' in cells[0]:
            cap_row = cells[1:]
    by_year = []
    for i, y in enumerate(years):
        try:
            util = float(util_row[i]) if util_row and i < len(util_row) else None
        except ValueError:
            util = None
        try:
            cap = float(cap_row[i]) if cap_row and i < len(cap_row) else None
        except ValueError:
            cap = None
        if util is not None or cap is not None:
            by_year.append({'year': y, 'util': util, 'capFactor': cap})
    return {
        'asOf':      datetime.now(timezone.utc).strftime('%Y-%m-%d'),
        'source':    'KHNP 이용률·가동률',
        'sourceUrl': url,
        'byYear':    by_year,
    }


def fetch_khnp_trips():
    """Annual unplanned-trip counts from KHNP '불시정지' page.

    Source: https://npp.khnp.co.kr/ON004004002001004
    Three stacked tables (2001-09, 2010-19, 2020-29). Each row: year |
    operating reactors | trip count | trips per reactor.
    Returns: { asOf, byYear: [{year, reactors, trips, tripsPerReactor}, ...] }
    """
    url = 'https://npp.khnp.co.kr/ON004004002001004'
    html = fetch(url)
    tables = re.findall(r'<table[^>]*>.*?</table>', html, re.DOTALL)
    by_year = []
    for tbl in tables[:3]:
        th_all = [re.sub(r'<[^>]+>', '', t).strip()
                  for t in re.findall(r'<th[^>]*>(.*?)</th>', tbl, re.DOTALL)]
        years = [int(t) for t in th_all if t.isdigit() and len(t) == 4]
        if not years:
            continue
        rows = re.findall(r'<tr[^>]*>(.*?)</tr>', tbl, re.DOTALL)
        reactors_row = trips_row = per_row = None
        for r in rows:
            cells = [re.sub(r'<[^>]+>', '', c).strip()
                     for c in re.findall(r'<t[dh][^>]*>(.*?)</t[dh]>', r, re.DOTALL)]
            if not cells:
                continue
            head = cells[0]
            data = cells[1:]
            # Order matters: '기당 불시 정지 건수' contains '불시 정지 건수' as a
            # substring, so check the per-reactor row FIRST.
            if '운전기수' in head:
                reactors_row = data
            elif '기당' in head:
                per_row = data
            elif '불시 정지 건수' in head:
                trips_row = data
        for i, y in enumerate(years):
            def _i(arr):
                if not arr or i >= len(arr) or arr[i] in ('', '-'):
                    return None
                try: return int(arr[i])
                except ValueError:
                    try: return float(arr[i])
                    except ValueError: return None
            entry = {
                'year':            y,
                'reactors':        _i(reactors_row),
                'trips':           _i(trips_row),
                'tripsPerReactor': _i(per_row),
            }
            if entry['reactors'] is not None or entry['trips'] is not None:
                by_year.append(entry)
    return {
        'asOf':      datetime.now(timezone.utc).strftime('%Y-%m-%d'),
        'source':    'KHNP 불시정지',
        'sourceUrl': url,
        'byYear':    by_year,
    }


def fetch_ember_monthly():
    """Monthly Korean electricity by source from Ember.

    Source: files.ember-energy.org public CSV (the same dataset behind
    their Monthly Electricity Data explorer). Free, no API key needed.
    The file is large (~70 MB) so we cache the parsed Korea-nuclear rows
    in data/korea_ember_monthly.json and only re-download weekly.
    """
    import csv
    import io as _io

    cache_path = os.path.join(os.path.dirname(OUT_PATH), 'korea_ember_monthly.json')

    def _enrich(out):
        """Add utilizationPct + capacityGw to every series row.
        Idempotent — recomputes on every load so capacity-table edits flow through.
        """
        for r in (out.get('series') or []):
            d = r.get('date'); twh = r.get('twh')
            if not d or twh is None: continue
            cap_gw = korea_nuc_capacity_gw(d)
            y, m = int(d[:4]), int(d[5:7])
            hours = calendar.monthrange(y, m)[1] * 24
            max_twh = cap_gw * hours / 1000.0  # GW × h = GWh; /1000 → TWh
            if max_twh > 0:
                r['utilizationPct'] = round(twh / max_twh * 100, 1)
                r['capacityGw'] = cap_gw
        return out

    # Refresh if cache missing or > 7 days old
    refresh = True
    if os.path.exists(cache_path):
        try:
            with open(cache_path, 'r', encoding='utf-8') as f:
                cached = json.load(f)
            cached_at = cached.get('cachedAt')
            if cached_at:
                from datetime import datetime as _dt
                age_days = (datetime.now(timezone.utc) - _dt.fromisoformat(cached_at)).days
                if age_days < 7:
                    return _enrich(cached)
        except Exception:
            pass

    if refresh:
        url = 'https://files.ember-energy.org/public-downloads/monthly_full_release_long_format.csv'
        try:
            text = fetch(url, timeout=60)
        except Exception as e:
            print(f'Ember monthly fetch failed: {e}', file=sys.stderr)
            # Return cached even if stale
            if os.path.exists(cache_path):
                with open(cache_path, 'r', encoding='utf-8') as f:
                    return json.load(f)
            return None
        rows = []
        reader = csv.DictReader(_io.StringIO(text))
        for r in reader:
            if r.get('Area') != 'South Korea': continue
            if r.get('Variable') != 'Nuclear': continue
            if r.get('Subcategory') != 'Fuel': continue
            unit = r.get('Unit')
            if unit not in ('TWh', '%'): continue
            try:
                val = float(r['Value']) if r.get('Value') else None
            except ValueError:
                val = None
            if val is None: continue
            rows.append({'date': r['Date'], 'unit': unit, 'value': val})

        # Pivot: date → {twh, share_pct}
        by_date = {}
        for r in rows:
            d = by_date.setdefault(r['date'], {})
            if r['unit'] == 'TWh': d['twh'] = r['value']
            elif r['unit'] == '%':  d['sharePct'] = r['value']
        series = sorted([{
            'date':     k,
            'twh':      v.get('twh'),
            'sharePct': v.get('sharePct'),
        } for k, v in by_date.items()], key=lambda x: x['date'])

        out = {
            'cachedAt': datetime.now(timezone.utc).isoformat(timespec='seconds'),
            'source':   'Ember Monthly Electricity Data',
            'sourceUrl':url,
            'series':   series,
        }
        _enrich(out)
        with open(cache_path, 'w', encoding='utf-8') as f:
            json.dump(out, f, indent=2)
        return out


# ────────────────────────────────────────────────────────────────────
# DART (OpenDART) — KOGAS quarterly filings + sectoral sales volumes
#
# Free API key required (env var OPENDART_KEY). Without it this fetcher
# returns None gracefully — existing data flow unaffected.
#
# KOGAS corp metadata (verified via corpCode.xml dump):
#   corp_code  = 00261285
#   corp_name  = 한국가스공사 (Korea Gas Corporation)
#   stock_code = 036460
#
# What we extract from the latest quarterly/half-year/annual report:
#   1. Annual sales volume by sector (도시가스용 / 발전용 / 합계) — historical
#      time series from 1987 onwards, in thousand tons.
#   2. LNG procurement origins (e.g., Qatar, Australia, US).
#   3. Forward-contracted volumes by tariff type (3 most recent years).
#   4. Latest period summary (revenue, op profit) from fnlttSinglAcntAll.
# ────────────────────────────────────────────────────────────────────
DART_KOGAS_CORP_CODE = '00261285'
DART_API_BASE = 'https://opendart.fss.or.kr/api'


def _dart_get(url, timeout=30):
    """GET a DART JSON endpoint and parse — raises on non-200."""
    text = fetch(url, timeout=timeout)
    return json.loads(text)


def _dart_get_zip(url, timeout=600):
    """Download a DART ZIP archive (corpCode.xml or document.xml) and return raw bytes.
    DART throttles uploads to ~13 KB/s — a 10MB document.xml can take 10+ min.
    Falls back to curl with retries if Python streaming truncates."""
    headers = {'User-Agent': UA}
    if _USE_REQUESTS:
        try:
            r = _requests.get(url, timeout=timeout, stream=True, headers=headers)
            r.raise_for_status()
            # Read in chunks so the connection stays alive on slow streams.
            chunks = []
            for chunk in r.iter_content(chunk_size=64*1024):
                if chunk:
                    chunks.append(chunk)
            content = b''.join(chunks)
            cl = r.headers.get('Content-Length')
            if not cl or int(cl) == len(content):
                return content
            print(f'DART zip: requests got {len(content)}/{cl} bytes — retrying via curl',
                  file=sys.stderr)
        except Exception as e:
            print(f'DART zip via requests failed: {e} — retrying via curl', file=sys.stderr)

    # Curl fallback with retries — DART occasionally drops the connection.
    import subprocess, tempfile
    with tempfile.NamedTemporaryFile(delete=False, suffix='.zip') as tf:
        tmp_path = tf.name
    try:
        subprocess.run(
            ['curl', '--http1.1', '-fsS', '-A', UA,
             '--max-time', str(timeout),
             '--connect-timeout', '30',
             '--retry', '3', '--retry-delay', '5',
             '--speed-time', '120', '--speed-limit', '300',
             '-o', tmp_path, url],
            check=True, timeout=timeout + 60,
        )
        with open(tmp_path, 'rb') as f:
            return f.read()
    finally:
        try: os.unlink(tmp_path)
        except: pass


def _dart_parse_period_report_xml(xml_text):
    """Parse a KOGAS period-report main XML. Returns dict with sectoral data.

    The report is a custom KICS XBRL-like document. We extract data from
    the embedded <TABLE> elements using regex (BeautifulSoup not available
    in the runtime). Three target tables:

      A) Annual sales volume by sector (multi-year, in kilo-tons):
         signature: contains '도시가스용' AND '발전용' AND '평균증가율'.
      B) Forward contracted volumes by tariff type (3 yrs × 3 cats):
         signature: '발전용 평균요금제' AND '도시가스용'.
      C) LNG procurement origins (single row — Qatar, Australia, US, ...):
         signature: 'LNG' AND ('카타르' OR '호주') AND short table.
    """
    import re
    out = {}

    tables = re.findall(r'<TABLE[\s\S]*?</TABLE>', xml_text)

    def _cells(table):
        """Strip XML tags from a table, collapse whitespace, split on '|'."""
        clean = re.sub(r'<[^>]+>', '|', table)
        clean = re.sub(r'\|+', '|', clean)
        clean = re.sub(r'\s+', ' ', clean).strip('| ').strip()
        return [c.strip() for c in clean.split('|')]

    def _to_int(s):
        s = re.sub(r'\([^)]*\)', '', s).replace(',', '').strip()
        try:    return int(float(s))
        except: return None

    SECTOR_SEPARATORS = {'도시가스용','발전용','발전용 평균요금제','발전용 개별요금제',
                          '합 계','합계','구 분','구분','용 도','용도'}

    def _row_after(cells, sector_kr, max_take=None):
        """Return ints from cells after sector_kr until next sector header.
        Strips parenthetical growth rates and decimals (CAGR), keeps only
        non-negative integer-looking numbers."""
        try:
            idx = cells.index(sector_kr)
        except ValueError:
            return None
        row_text_parts = []
        for c in cells[idx+1:]:
            if c in SECTOR_SEPARATORS:
                break
            row_text_parts.append(c)
        row_text = ' '.join(row_text_parts)
        # 1) Drop everything inside parens (growth rates like (-6.9), (2.7))
        row_text = re.sub(r'\([^)]*\)', ' ', row_text)
        # 2) Drop standalone decimal numbers like CAGR "15.7%" or "8.4%"
        row_text = re.sub(r'\b\d+\.\d+\s*%?', ' ', row_text)
        # 3) Now extract integer tokens (with optional thousands separator)
        ints = []
        for m in re.finditer(r'-?\d{1,3}(?:,\d{3})+|\d+', row_text):
            n = _to_int(m.group(0))
            if n is None or n < 0 or n > 1_000_000:
                continue
            ints.append(n)
            if max_take and len(ints) >= max_take:
                break
        return ints

    # ── Table A: annual sectoral sales volume ──────────────────────────
    for t in tables:
        if not ('도시가스용' in t and '발전용' in t and '평균증가율' in t):
            continue
        if len(t) > 12000:  # skip huge boilerplate tables
            continue
        cells = _cells(t)
        # Years are 4-digit ints between 1987 and current+1.
        years = []
        for c in cells:
            m = re.fullmatch(r'(19\d\d|20\d\d)', c)
            if m:
                y = int(m.group(1))
                if 1987 <= y <= datetime.now().year + 1 and y not in years:
                    years.append(y)
        if not years:
            continue

        rows_data = {}
        for sector_kr, key in [('도시가스용','cityGas_kt'),
                                ('발전용','power_kt'),
                                ('합 계','total_kt'),
                                ('합계','total_kt')]:
            ints = _row_after(cells, sector_kr, max_take=len(years) + 1)
            if ints and len(ints) >= len(years):
                rows_data[key] = ints[:len(years)]

        if rows_data:
            entries = []
            for i, y in enumerate(years):
                e = {'year': y}
                for k, vals in rows_data.items():
                    e[k] = vals[i]
                entries.append(e)
            out['sectorVolumesByYear'] = entries
            break

    # ── Table B: forward-contracted volumes by tariff type ─────────────
    for t in tables:
        if not ('발전용 평균요금제' in t and '도시가스용' in t):
            continue
        if len(t) > 6000:
            continue
        cells = _cells(t)
        # Years like '25년, '24년, '23년 → ints 2025/2024/2023
        years = []
        for c in cells:
            m = re.search(r"'(\d{2})년", c)
            if m:
                y = 2000 + int(m.group(1))
                if y not in years:
                    years.append(y)
        if not years or len(years) < 2:
            continue
        # Re-use the row helper (allow decimals here — values may be 962.5)
        def _row_floats(cells, sector_kr, n_years):
            try:
                idx = cells.index(sector_kr)
            except ValueError:
                return None
            text_parts = []
            for c in cells[idx+1:]:
                if c in SECTOR_SEPARATORS:
                    break
                text_parts.append(c)
            text = ' '.join(text_parts)
            text = re.sub(r'\([^)]*\)', ' ', text)
            nums = re.findall(r'-?\d{1,3}(?:,\d{3})+(?:\.\d+)?|\d+(?:\.\d+)?', text)
            vals = []
            for n in nums:
                try:    vals.append(float(n.replace(',', '')))
                except: pass
                if len(vals) >= n_years:
                    break
            return vals if len(vals) >= n_years else None

        rows_data = {}
        for sector_kr, key in [
            ('발전용 평균요금제', 'power_avg_tariff_kt'),
            ('발전용 개별요금제', 'power_individual_tariff_kt'),
            ('도시가스용',         'cityGas_kt'),
            ('합 계',              'total_kt'),
            ('합계',                'total_kt'),
        ]:
            vals = _row_floats(cells, sector_kr, len(years))
            if vals:
                rows_data[key] = vals[:len(years)]
        if rows_data:
            out['contractedVolumes'] = [
                {'year': y, **{k: vals[i] for k, vals in rows_data.items()}}
                for i, y in enumerate(years)
            ]
            break

    # ── Table C: LNG procurement origins ───────────────────────────────
    for t in tables:
        if 'LNG' in t and ('카타르' in t or '호주' in t) and len(t) < 3500:
            cells = _cells(t)
            for c in cells:
                if '카타르' in c or '호주' in c:
                    raw = re.sub(r'\s*등\s*$', '', c)
                    parts = re.split(r'[,、·\s]+', raw)
                    origins_kr = [p for p in parts if p and len(p) <= 6]
                    if origins_kr:
                        # Korean → English mapping for display
                        kr2en = {
                            '카타르':'Qatar','호주':'Australia','미국':'United States',
                            '오만':'Oman','말레이시아':'Malaysia','인도네시아':'Indonesia',
                            '러시아':'Russia','나이지리아':'Nigeria','브루나이':'Brunei',
                            '예멘':'Yemen','앙골라':'Angola','이집트':'Egypt',
                            '트리니다드':'Trinidad','UAE':'UAE','파푸아뉴기니':'Papua New Guinea',
                        }
                        origins_en = [kr2en.get(k, k) for k in origins_kr]
                        out['lngOrigins'] = origins_en
                        out['lngOriginsKr'] = origins_kr
                        break
            if 'lngOrigins' in out:
                break

    # ── Long-term LNG supply contracts ──────────────────────────────────
    # Look for the table with '계약기간' + '연간계약물량' + '천톤' headers.
    for t in tables:
        if ('연간계약물량' in t and '천톤' in t and '계약기간' in t
                and len(t) < 50000 and len(t) > 5000):
            contracts = _parse_supply_contracts(t)
            if contracts and len(contracts) >= 5:
                out['supplyContracts'] = contracts
                break

    # ── Annual send-out by terminal (tons) ──────────────────────────────
    for t in tables:
        if ('평택기지' in t and '인천기지' in t and '한국가스공사' in t
                and '톤' in t and len(t) < 8000):
            sendout = _parse_terminal_sendout(t)
            if sendout:
                out['terminalSendout'] = sendout
                break

    return out


def _dart_decode_disclosure(buf):
    """Decode a KOGAS DART disclosure HTML, ignoring the (often-wrong) meta
    charset declaration. Newer (2022+) disclosures declare 'euc-kr' but
    actually contain UTF-8 bytes. Detect by looking for known Korean tech
    terms in the candidate decoded text."""
    KEYWORDS = ('도시가스', '발전용', '잠정', '실적')
    # Try UTF-8 first — most disclosures (incl. 2022+) are UTF-8 despite meta tag
    try:
        text = buf.decode('utf-8')
        if any(k in text for k in KEYWORDS):
            return text
    except UnicodeDecodeError:
        pass
    # Fall back to cp949/EUC-KR for genuinely old (pre-2022) filings
    text = buf.decode('cp949', errors='replace')
    return text


def _parse_kogas_monthly_disclosure(html):
    """Parse one KOGAS '영업(잠정)실적(공정공시)' disclosure HTML.

    Returns dict with monthly sales volume by sector (kt) for the
    period indicated by 당기실적, plus prev-month + same-month-last-year
    reference values (for diff-checks).
    """
    import re, html as _html
    text = re.sub(r'<style[\s\S]*?</style>', '', html, flags=re.I)
    text = re.sub(r'<script[\s\S]*?</script>', '', text, flags=re.I)
    # Decode HTML entities (&nbsp; etc.) BEFORE regex extraction so the
    # sector labels match cleanly. Older disclosures wrap label spaces in
    # &nbsp; (e.g. '기 &nbsp;타' = '기 타').
    text = _html.unescape(text)
    flat = re.sub(r'<[^>]+>', '|', text)
    flat = re.sub(r'\|+', '|', flat)
    flat = re.sub(r'\s+', ' ', flat)

    # ── Period: find the 당기실적 date range YYYY-MM-DD ~ YYYY-MM-DD ──────
    # In the cleaned text, after '당기실적' the next two YYYY-MM-DD pairs are
    # current-period start/end.
    period_year = None
    period_month = None
    m_dr = re.search(r'당기실적[^\d]*(\d{4})-(\d{2})-(\d{2})[^\d]*~[^\d]*(\d{4})-(\d{2})-(\d{2})', flat)
    if m_dr:
        period_year = int(m_dr.group(1))
        period_month = int(m_dr.group(2))
    else:
        # Fallback: pull '20'YY월' from the volume table header area.
        m_yy = re.search(r"\('?(\d{2})\.(\d{1,2})월\)", flat)
        if m_yy:
            period_year = 2000 + int(m_yy.group(1))
            period_month = int(m_yy.group(2))

    # ── Sectoral volumes ────────────────────────────────────────────────
    def _row_nums(label_alts):
        """Find any of the label alts in the cleaned text and return numeric
        tokens that follow, until the next sector keyword."""
        next_labels = ['도시가스용','발전용','기 타','기타','총 계','총계','2.','정보제공']
        for label in label_alts:
            idx = flat.find(label)
            if idx < 0:
                continue
            # carve out the row text (next 250 chars)
            tail = flat[idx + len(label):idx + len(label) + 350]
            # stop at next sector header
            stops = [tail.find(nl) for nl in next_labels if nl != label and tail.find(nl) > 0]
            if stops:
                tail = tail[:min(stops)]
            # extract all numbers
            nums = re.findall(r'-?\d{1,3}(?:,\d{3})+|\d+(?:\.\d+)?', tail)
            cleaned = [float(n.replace(',', '')) for n in nums if n != '-']
            return cleaned
        return []

    city_nums  = _row_nums(['도시가스용'])
    power_nums = _row_nums(['발전용'])
    total_nums = _row_nums(['총 계', '총계', '합 계', '합계'])

    def _first_int(arr):
        return int(arr[0]) if arr else None
    def _yoy_int(arr):
        # arr layout: [curr, prev, MoM%, prevYr, YoY%]  (5 nums)
        return int(arr[3]) if len(arr) >= 4 else None

    if not period_year or not period_month:
        return None
    if not (city_nums or power_nums or total_nums):
        return None

    return {
        'date':         f'{period_year}-{period_month:02d}-01',
        'cityGas_kt':   _first_int(city_nums),
        'power_kt':     _first_int(power_nums),
        'total_kt':     _first_int(total_nums),
        'cityGas_kt_yoy':  _yoy_int(city_nums),
        'power_kt_yoy':    _yoy_int(power_nums),
        'total_kt_yoy':    _yoy_int(total_nums),
    }


def fetch_dart_kogas_monthly():
    """Pull every KOGAS '영업(잠정)실적(공정공시)' disclosure from DART and
    build a monthly sectoral sales time series.

    KOGAS files one of these every month with monthly volumes for City-gas
    and Power. Each disclosure is ~2.5KB HTML, so ingesting all backfill
    is fast (~75 disclosures × 2.5KB = 200KB).

    Caches by rcept_no in data/dart_kogas_monthly.json. Skips disclosures
    already parsed; only fetches new ones (and re-fetches the most recent
    3 in case of revisions).
    """
    api_key = os.environ.get('OPENDART_KEY')
    if not api_key:
        return None

    cache_path = os.path.join(os.path.dirname(OUT_PATH), 'dart_kogas_monthly.json')
    cached = {'series': [], 'parsedRcepts': []}
    if os.path.exists(cache_path):
        try:
            with open(cache_path, 'r', encoding='utf-8') as f:
                cached = json.load(f)
        except Exception:
            cached = {'series': [], 'parsedRcepts': []}

    parsed_set = set(cached.get('parsedRcepts') or [])
    series_by_date = {r['date']: r for r in (cached.get('series') or [])}

    # ── Enumerate all KOGAS 영업(잠정)실적 disclosures since 2020 ────────
    today = datetime.now(timezone.utc)
    all_filings = []
    for yr in range(2020, today.year + 1):
        url = (f'{DART_API_BASE}/list.json'
               f'?crtfc_key={api_key}&corp_code={DART_KOGAS_CORP_CODE}'
               f'&bgn_de={yr}0101&end_de={yr}1231&page_count=100')
        try:
            data = _dart_get(url, timeout=30)
        except Exception as e:
            print(f'DART monthly list {yr} failed: {e}', file=sys.stderr)
            continue
        if data.get('status') != '000':
            continue
        for f in (data.get('list') or []):
            nm = (f.get('report_nm') or '').strip()
            # Match: 영업(잠정)실적(공정공시) — but skip 연결재무제표기준 (consolidated FS, quarterly)
            if '영업(잠정)실적' in nm and '연결' not in nm:
                all_filings.append({
                    'rcept_no': f['rcept_no'],
                    'rcept_dt': f['rcept_dt'],
                    'report_nm': nm,
                })

    # ── Determine which to fetch: new + re-fetch latest 3 ────────────────
    all_rcepts = {f['rcept_no'] for f in all_filings}
    # Sort by rcept_no descending (most recent first); rcept_no is sortable.
    sorted_filings = sorted(all_filings, key=lambda f: f['rcept_no'], reverse=True)
    refetch_recent = {f['rcept_no'] for f in sorted_filings[:3]}
    todo = [f for f in sorted_filings
            if f['rcept_no'] not in parsed_set or f['rcept_no'] in refetch_recent]

    print(f'DART monthly: {len(all_filings)} disclosures total, '
          f'{len(parsed_set)} cached, {len(todo)} to fetch', file=sys.stderr)

    # ── Fetch + parse each ──────────────────────────────────────────────
    import zipfile, io as _io
    fetched = 0
    for f in todo:
        rno = f['rcept_no']
        url = f'{DART_API_BASE}/document.xml?crtfc_key={api_key}&rcept_no={rno}'
        try:
            buf = _dart_get_zip(url, timeout=120)
            with zipfile.ZipFile(_io.BytesIO(buf)) as z:
                with z.open(z.namelist()[0]) as fp:
                    raw = fp.read()
            html = _dart_decode_disclosure(raw)
            parsed = _parse_kogas_monthly_disclosure(html)
            if parsed:
                parsed['rcept_no'] = rno
                parsed['rcept_dt'] = f['rcept_dt']
                series_by_date[parsed['date']] = parsed
                parsed_set.add(rno)
                fetched += 1
        except Exception as e:
            print(f'DART monthly disc {rno}: {e}', file=sys.stderr)

    # ── Persist cache ────────────────────────────────────────────────────
    series = sorted(series_by_date.values(), key=lambda r: r['date'])
    out = {
        'cachedAt': today.isoformat(timespec='seconds'),
        'source':   'OpenDART · KOGAS 영업(잠정)실적(공정공시) monthly disclosures',
        'sourceUrl':'https://opendart.fss.or.kr',
        'parsedRcepts': sorted(parsed_set),
        'series':   series,
    }
    with open(cache_path, 'w', encoding='utf-8') as f:
        json.dump(out, f, indent=2, ensure_ascii=False)
    print(f'DART monthly: {len(series)} months in series '
          f'(fetched {fetched} new this run)', file=sys.stderr)
    return out


def _dart_extract_quarterly_inventory(api_key, corp_code, year, reprt_code):
    """Pull a single quarterly snapshot from fnlttSinglAcntAll.
    Returns dict with end_date, inventory_krw, cogs_krw, sales_krw — or None."""
    url = (f'{DART_API_BASE}/fnlttSinglAcntAll.json'
           f'?crtfc_key={api_key}&corp_code={corp_code}'
           f'&bsns_year={year}&reprt_code={reprt_code}&fs_div=CFS')
    try:
        data = _dart_get(url, timeout=30)
    except Exception as e:
        return None
    if data.get('status') != '000':
        return None

    items = data.get('list') or []
    out = {'year': year, 'reprt_code': reprt_code}

    # Period-end date by report code
    end_month = {'11013': 3, '11012': 6, '11014': 9, '11011': 12}.get(reprt_code)
    if end_month:
        # Last day of the period quarter
        if end_month == 3:    out['end_date'] = f'{year}-03-31'
        elif end_month == 6:  out['end_date'] = f'{year}-06-30'
        elif end_month == 9:  out['end_date'] = f'{year}-09-30'
        else:                 out['end_date'] = f'{year}-12-31'
        out['quarter'] = end_month // 3   # 1, 2, 3, 4

    def _amt(item):
        try:    return int(item.get('thstrm_amount', '').replace(',', ''))
        except (ValueError, AttributeError): return None

    for r in items:
        nm = r.get('account_nm', '') or ''
        sj = r.get('sj_nm', '') or ''
        if sj == '재무상태표' and nm == '재고자산':
            out['inventory_krw'] = _amt(r)
        elif sj == '포괄손익계산서' and nm == '매출원가':
            out['cogs_krw'] = _amt(r)
        elif sj == '포괄손익계산서' and nm == '재화의 판매로 인한 수익':
            out['sales_krw'] = _amt(r)

    if out.get('inventory_krw') is None:
        return None
    return out


def _parse_supply_contracts(table_xml):
    """Parse KOGAS long-term LNG supply contracts table.

    Layout (alternating empty cells between data):
      [name] [_] [period like '2008~2028'] [_] [volume like '2,000.0'] [_]...

    We strip empties first then scan triplets.
    """
    import re
    text = re.sub(r'<[^>]+>', '|', table_xml)
    text = re.sub(r'\|+', '|', text)
    text = re.sub(r'\s+', ' ', text)
    raw = [c.strip() for c in text.split('|')]
    # Drop blanks AND header labels we don't want as counterparties
    HEADER_TOKENS = {'계약기간','연간계약물량(단위: 천톤)','약정의 유형',
                      '매입약정','재고자산 구매계약','거래상대방'}
    cells = [c for c in raw if c and c not in HEADER_TOKENS]

    contracts = []
    i = 0
    while i + 2 < len(cells):
        name = cells[i]
        period = cells[i + 1]
        vol_str = cells[i + 2]
        m_period = re.match(r'(\d{4})\s*~\s*(\d{4})$', period)
        m_vol = re.match(r'^([\d,]+(?:\.\d+)?)$', vol_str)
        if m_period and m_vol and len(name) >= 3 and not name[0].isdigit():
            contracts.append({
                'counterparty': name,
                'startYear':    int(m_period.group(1)),
                'endYear':      int(m_period.group(2)),
                'annual_kt':    float(m_vol.group(1).replace(',', '')),
            })
            i += 3
        else:
            i += 1
    # Dedupe (table appears twice in report)
    seen = set()
    unique = []
    for c in contracts:
        key = (c['counterparty'], c['startYear'], c['endYear'])
        if key not in seen:
            seen.add(key)
            unique.append(c)
    return unique


def _parse_terminal_sendout(table_xml):
    """Parse the 'Annual send-out by terminal (tons)' table — typically Table #39.
    Row format: terminal | YYYY current | YYYY-1 | YYYY-2 (3 yrs)."""
    import re
    text = re.sub(r'<[^>]+>', '|', table_xml)
    text = re.sub(r'\|+', '|', text)
    text = re.sub(r'\s+', ' ', text)
    cells = [c.strip() for c in text.split('|')]

    # Find years like '25년 / '24년 / '23년
    years = []
    for c in cells:
        m = re.match(r"'(\d{2})년", c)
        if m and 2000 + int(m.group(1)) not in years:
            years.append(2000 + int(m.group(1)))
        if len(years) >= 3: break
    if len(years) < 2:
        return None

    TERMINALS = {'평택기지':'pyeongtaek', '인천기지':'incheon',
                  '통영기지':'tongyeong', '삼척기지':'samcheok',
                  '제주기지':'jeju'}
    rows = []
    for kr_name, en_key in TERMINALS.items():
        try:
            idx = cells.index(kr_name)
        except ValueError:
            continue
        nums = []
        for c in cells[idx+1:idx+1+2*len(years)]:
            m = re.match(r'^([\d,]+)$', c.replace(' ', ''))
            if m: nums.append(int(m.group(1).replace(',', '')))
            if len(nums) >= len(years): break
        if len(nums) >= len(years):
            rows.append({'terminal': en_key, 'terminal_kr': kr_name,
                          **{str(y): nums[i] for i, y in enumerate(years)}})
    return rows if rows else None


def fetch_dart_kogas():
    """Pull KOGAS sectoral data from OpenDART API.

    Requires environment variable OPENDART_KEY. Returns None if missing.
    Caches the parsed report in data/dart_kogas.json — only re-parses
    when a new period report (분기/반기/사업보고서) shows up in the listing.
    """
    api_key = os.environ.get('OPENDART_KEY')
    if not api_key:
        print('OPENDART_KEY not set — skipping DART fetch', file=sys.stderr)
        return None

    cache_path = os.path.join(os.path.dirname(OUT_PATH), 'dart_kogas.json')
    today = datetime.now(timezone.utc)

    # 1. List recent KOGAS filings (last 2 years).
    bgn = (today - timedelta(days=730)).strftime('%Y%m%d')
    end = today.strftime('%Y%m%d')
    list_url = (f'{DART_API_BASE}/list.json'
                f'?crtfc_key={api_key}&corp_code={DART_KOGAS_CORP_CODE}'
                f'&bgn_de={bgn}&end_de={end}&page_count=100')
    try:
        list_data = _dart_get(list_url, timeout=30)
    except Exception as e:
        print(f'DART list fetch failed: {e}', file=sys.stderr)
        return None

    if list_data.get('status') != '000':
        print(f"DART list error: {list_data.get('message')}", file=sys.stderr)
        return None
    filings = list_data.get('list') or []

    # 2. Find latest period report (분기/반기/사업보고서 — full-form business report).
    latest_period = None
    for f in filings:
        nm = f.get('report_nm', '') or ''
        if '분기보고서' in nm or '반기보고서' in nm or '사업보고서' in nm:
            latest_period = f
            break

    # 3. Cache check — skip parse if same rcept_no.
    cached = None
    if os.path.exists(cache_path):
        try:
            with open(cache_path, 'r', encoding='utf-8') as f:
                cached = json.load(f)
        except Exception:
            cached = None

    if (cached and latest_period
            and cached.get('latestRceptNo') == latest_period.get('rcept_no')):
        cached['filings'] = [
            {'rcept_no': f['rcept_no'], 'rcept_dt': f['rcept_dt'],
             'report_nm': (f['report_nm'] or '').strip()}
            for f in filings[:30]
        ]
        cached['asOf'] = today.isoformat(timespec='seconds')
        with open(cache_path, 'w', encoding='utf-8') as f:
            json.dump(cached, f, indent=2, ensure_ascii=False)
        return cached

    # 4. Download + parse latest period report main XML.
    sector_data = {}
    if latest_period:
        rcept_no = latest_period['rcept_no']
        doc_url = f'{DART_API_BASE}/document.xml?crtfc_key={api_key}&rcept_no={rcept_no}'
        try:
            buf = _dart_get_zip(doc_url, timeout=300)
            import zipfile, io as _io
            with zipfile.ZipFile(_io.BytesIO(buf)) as z:
                with z.open(z.namelist()[0]) as f:
                    xml_text = f.read().decode('utf-8', errors='replace')
            sector_data = _dart_parse_period_report_xml(xml_text)
            print(f"DART parsed report {rcept_no}: "
                  f"annual={len(sector_data.get('sectorVolumesByYear') or [])}y, "
                  f"contracted={len(sector_data.get('contractedVolumes') or [])}y, "
                  f"origins={sector_data.get('lngOrigins') or []}",
                  file=sys.stderr)
        except Exception as e:
            print(f'DART document.xml parse failed: {e}', file=sys.stderr)

    out = {
        'asOf':              today.isoformat(timespec='seconds'),
        'corpCode':          DART_KOGAS_CORP_CODE,
        'corpName':          '한국가스공사',
        'corpNameEn':        'Korea Gas Corporation',
        'stockCode':         '036460',
        'latestRceptNo':     latest_period['rcept_no'] if latest_period else None,
        'latestReportName':  (latest_period['report_nm'].strip() if latest_period else None),
        'latestReportDate':  latest_period['rcept_dt'] if latest_period else None,
        'filings': [
            {'rcept_no': f['rcept_no'], 'rcept_dt': f['rcept_dt'],
             'report_nm': (f['report_nm'] or '').strip()}
            for f in filings[:30]
        ],
        'sectorVolumesByYear': sector_data.get('sectorVolumesByYear'),
        'contractedVolumes':   sector_data.get('contractedVolumes'),
        'lngOrigins':          sector_data.get('lngOrigins'),
        'lngOriginsKr':        sector_data.get('lngOriginsKr'),
        'supplyContracts':     sector_data.get('supplyContracts'),
        'terminalSendout':     sector_data.get('terminalSendout'),
        'source':              'OpenDART (Financial Supervisory Service)',
        'sourceUrl':           'https://opendart.fss.or.kr',
    }

    with open(cache_path, 'w', encoding='utf-8') as f:
        json.dump(out, f, indent=2, ensure_ascii=False)
    return out


def fetch_dart_kogas_quarterly_inventory():
    """Pull KOGAS quarterly inventory snapshots (KRW) from DART
    fnlttSinglAcntAll. Loops 2020 → current year × 4 reprt_codes.

    Cached in data/dart_kogas_quarterly.json, keyed by (year, reprt_code).
    Only re-fetches missing snapshots. Cheap — each call returns ~300 lines
    of JSON (~30KB), 5 yrs × 4 quarters = 20 calls = ~600KB total.

    Computes implied physical inventory in kt using the avg cost basis
    derived from cumulative COGS / cumulative monthly sales.
    """
    api_key = os.environ.get('OPENDART_KEY')
    if not api_key:
        return None

    cache_path = os.path.join(os.path.dirname(OUT_PATH), 'dart_kogas_quarterly.json')
    cached = {'snapshots': []}
    if os.path.exists(cache_path):
        try:
            with open(cache_path, 'r', encoding='utf-8') as f:
                cached = json.load(f)
        except Exception:
            cached = {'snapshots': []}

    # Index by (year, reprt_code)
    snaps_by_key = {(s['year'], s['reprt_code']): s for s in (cached.get('snapshots') or [])}

    today = datetime.now(timezone.utc)
    REPRT_CODES = ['11013', '11012', '11014', '11011']  # Q1, H1, Q3, Annual

    fetched = 0
    for year in range(2020, today.year + 1):
        for rc in REPRT_CODES:
            key = (year, rc)
            # Skip future periods (annual 2026 won't exist until April 2027)
            month = {'11013': 5, '11012': 8, '11014': 11, '11011': 4}[rc]
            year_avail = year if rc != '11011' else year + 1
            avail_date = datetime(year_avail, month, 1, tzinfo=timezone.utc)
            if today < avail_date:
                continue
            # Re-fetch the most recent (might be revised)
            recent = (year == today.year and rc == REPRT_CODES[0])
            if key in snaps_by_key and not recent:
                continue
            snap = _dart_extract_quarterly_inventory(api_key, DART_KOGAS_CORP_CODE, year, rc)
            if snap:
                snaps_by_key[key] = snap
                fetched += 1
                time.sleep(0.5)  # be polite to DART API

    snapshots = sorted(snaps_by_key.values(), key=lambda s: s['end_date'])

    # Compute implied physical inventory using cumulative cost basis.
    # Avg cost basis (KRW/kt) = cumulative COGS / cumulative monthly sales (kt).
    # We use the monthly-sales series if available (already cached locally).
    monthly_path = os.path.join(os.path.dirname(OUT_PATH), 'dart_kogas_monthly.json')
    monthly_series = []
    if os.path.exists(monthly_path):
        try:
            with open(monthly_path, 'r', encoding='utf-8') as f:
                monthly_series = json.load(f).get('series') or []
        except Exception:
            pass

    # Cumulative kt sold by year (sum of monthly totals up to and including month M)
    def cum_kt(end_date):
        y = int(end_date[:4]); m = int(end_date[5:7])
        return sum((r.get('total_kt') or 0)
                   for r in monthly_series
                   if r['date'][:4] == str(y) and int(r['date'][5:7]) <= m)

    # Derive a 'normalised' inventory kt using ANNUAL avg cost basis (more stable
    # than per-quarter cumulative ratios, which get distorted by term/spot mix).
    # For each quarter, look up the annual cost basis from the SAME year's
    # full-year COGS (from the Q4/annual snapshot) divided by full-year sales kt.
    annual_cost_basis = {}  # year → avg KRW/kt
    for s in snapshots:
        if s.get('reprt_code') != '11011':  # only annual reports
            continue
        cogs = s.get('cogs_krw')
        y = s['year']
        full_year_kt = sum((r.get('total_kt') or 0)
                            for r in monthly_series
                            if r['date'][:4] == str(y))
        if cogs and full_year_kt > 0:
            annual_cost_basis[y] = cogs / full_year_kt

    for s in snapshots:
        y = s['year']
        cost = annual_cost_basis.get(y) or annual_cost_basis.get(y - 1)
        if cost and cost > 0:
            inv = s.get('inventory_krw')
            if inv:
                s['implied_inventory_kt'] = round(inv / cost, 1)
                s['avg_cost_krw_per_kt'] = round(cost, 0)
        s['cumulative_sales_kt'] = cum_kt(s['end_date'])

    out = {
        'cachedAt':  today.isoformat(timespec='seconds'),
        'source':    'OpenDART · KOGAS quarterly fnlttSinglAcntAll',
        'sourceUrl': 'https://opendart.fss.or.kr',
        'snapshots': snapshots,
    }
    with open(cache_path, 'w', encoding='utf-8') as f:
        json.dump(out, f, indent=2, ensure_ascii=False)
    print(f'DART quarterly inventory: {len(snapshots)} snapshots '
          f'(fetched {fetched} new this run)', file=sys.stderr)
    return out


# ────────────────────────────────────────────────────────────────────
# Data.go.kr — Korean Public Data Portal (no API key required for file
# datasets; just two-step CSV download via session)
#
# Datasets used:
#   15129906  KOGAS monthly supply by use (city gas vs power) — tons,
#             monthly, Jan 2016 → present.
#   15052058  KOGAS power-sector natural-gas tariff (raw material cost +
#             supply cost) — KRW/Nm³ + KRW/GJ, monthly, Jan 2008 → present.
#
# Two-step protocol:
#   1) GET /tcs/dss/selectFileDataDownload.do?publicDataPk=...&fileDetailSn=1
#      → JSON metadata containing atchFileId
#   2) GET /cmm/cmm/fileDownload.do?atchFileId=...&fileDetailSn=1
#      → CSV bytes (cp949 encoded for Korean files)
# ────────────────────────────────────────────────────────────────────
DATA_GO_KR_BASE = 'https://www.data.go.kr'
KOGAS_POWER_TARIFF_URL = 'https://www.kogas.or.kr/site/koGas/1040402000000'


def _load_json_file(path):
    try:
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return None


def _parse_korean_float(s):
    try:
        return float(str(s).replace(',', '').strip())
    except (TypeError, ValueError):
        return None


def _iso_date_from_parts(y, m, d):
    return f'{int(y):04d}-{int(m):02d}-{int(d):02d}'


def _kogas_tariff_stale(end_iso, now=None):
    if not end_iso:
        return True
    now = now or datetime.now(timezone.utc)
    try:
        end_dt = datetime.fromisoformat(end_iso).replace(tzinfo=timezone.utc)
        return end_dt.date() < now.date()
    except Exception:
        return True


def fetch_kogas_current_power_tariff():
    """Pull current KOGAS power-sector tariff from the official KOGAS page.

    Source page exposes the effective period and the current KRW/GJ table:
      - direct-supplied power producers (100MW+)
      - general power producers / CHP

    The page can be slow or intermittently unreachable from non-Korean
    networks. On fetch/parse failure, return the last cached official value
    with a fetchError flag instead of inventing a tariff.
    """
    import html as _html
    cache_path = os.path.join(os.path.dirname(OUT_PATH), 'kogas_current_power_tariff.json')
    today = datetime.now(timezone.utc)
    cached = _load_json_file(cache_path)

    if cached and cached.get('cachedAt'):
        try:
            age_h = (today - datetime.fromisoformat(cached['cachedAt'])).total_seconds() / 3600
            if age_h < 12:
                cached['stale'] = _kogas_tariff_stale(cached.get('effectiveEnd'), today)
                return cached
        except Exception:
            pass

    try:
        raw = fetch(KOGAS_POWER_TARIFF_URL, timeout=30, retries=2)
        if isinstance(raw, bytes):
            raw = raw.decode('utf-8', errors='replace')
        text = _html.unescape(re.sub(r'<[^>]+>', ' ', raw))
        text = re.sub(r'\s+', ' ', text)

        date_re = re.search(
            r'(\d{4})\s*\.\s*(\d{1,2})\s*\.\s*(\d{1,2})\s*~\s*'
            r'(\d{4})\s*\.\s*(\d{1,2})\s*\.\s*(\d{1,2})',
            text
        )
        vals_re = re.search(
            r'원료비\s*([0-9,]+(?:\.[0-9]+)?)\s*([0-9,]+(?:\.[0-9]+)?)\s*'
            r'공급비\s*([0-9,]+(?:\.[0-9]+)?)\s*([0-9,]+(?:\.[0-9]+)?)\s*'
            r'합계\s*([0-9,]+(?:\.[0-9]+)?)\s*([0-9,]+(?:\.[0-9]+)?)',
            text
        )
        if not date_re or not vals_re:
            raise ValueError('effective period or tariff rows not found')

        start = _iso_date_from_parts(*date_re.groups()[:3])
        end = _iso_date_from_parts(*date_re.groups()[3:])
        vals = [_parse_korean_float(v) for v in vals_re.groups()]
        if any(v is None for v in vals):
            raise ValueError('tariff values not numeric')

        out = {
            'cachedAt': today.isoformat(timespec='seconds'),
            'source': 'KOGAS official · 발전용 천연가스 요금',
            'sourceUrl': KOGAS_POWER_TARIFF_URL,
            'unit': 'KRW/GJ',
            'effectiveStart': start,
            'effectiveEnd': end,
            'directPower': {
                'label': 'KOGAS direct-supplied power producers (100MW+)',
                'raw_krw_gj': vals[0],
                'supply_krw_gj': vals[2],
                'total_krw_gj': vals[4],
            },
            'generalPowerChp': {
                'label': 'General power producers / CHP',
                'raw_krw_gj': vals[1],
                'supply_krw_gj': vals[3],
                'total_krw_gj': vals[5],
            },
        }
        out['stale'] = _kogas_tariff_stale(out['effectiveEnd'], today)
        with open(cache_path, 'w', encoding='utf-8') as f:
            json.dump(out, f, indent=2, ensure_ascii=False)
        print(f"KOGAS current tariff: {out['effectiveStart']}–{out['effectiveEnd']} "
              f"direct={out['directPower']['total_krw_gj']} KRW/GJ",
              file=sys.stderr)
        return out
    except Exception as e:
        if cached:
            cached['fetchError'] = 'official page unreachable; using cached tariff'
            cached['stale'] = True
            print(f'KOGAS current tariff fetch failed, using cache: {e}', file=sys.stderr)
            return cached
        print(f'KOGAS current tariff fetch failed: {e}', file=sys.stderr)
        return None


def _data_go_kr_download(public_data_pk, file_detail_sn='1'):
    """Two-step Data.go.kr file fetch. Returns raw bytes of the CSV file
    or None on failure. No API key required."""
    if _USE_REQUESTS:
        sess = _requests.Session()
    else:
        return None  # urllib doesn't support session cleanly enough here
    sess.headers.update({
        'User-Agent': UA,
        'Referer': f'{DATA_GO_KR_BASE}/data/{public_data_pk}/fileData.do',
    })
    try:
        # 1) Metadata call
        meta_url = f'{DATA_GO_KR_BASE}/tcs/dss/selectFileDataDownload.do'
        r = sess.get(meta_url,
                     params={'publicDataPk': str(public_data_pk),
                              'fileDetailSn': str(file_detail_sn)},
                     timeout=30)
        r.raise_for_status()
        meta = r.json()
        atch_id = (meta.get('atchFileId')
                    or (meta.get('fileDataRegistVO') or {}).get('atchFileId'))
        if not atch_id:
            print(f'data.go.kr {public_data_pk}: no atchFileId', file=sys.stderr)
            return None
        # 2) File download
        dl_url = f'{DATA_GO_KR_BASE}/cmm/cmm/fileDownload.do'
        r = sess.get(dl_url,
                     params={'atchFileId': atch_id, 'fileDetailSn': file_detail_sn},
                     timeout=120)
        r.raise_for_status()
        return r.content
    except Exception as e:
        print(f'data.go.kr {public_data_pk} download failed: {e}', file=sys.stderr)
        return None


def _decode_korean_csv(buf):
    """Korean public-data CSVs may be utf-8-sig, cp949, or euc-kr."""
    for enc in ('utf-8-sig', 'utf-8', 'cp949', 'euc-kr'):
        try:
            return buf.decode(enc)
        except UnicodeDecodeError:
            continue
    return buf.decode('cp949', errors='replace')


def fetch_kogas_monthly_supply():
    """Pull KOGAS monthly supply by use (city gas + power) from Data.go.kr.
    No API key required. CSV columns: 연도, 월, 용도, 공급량 (tons).
    Coverage: Jan 2016 → present, refreshed quarterly.
    """
    import csv as _csv
    import io as _io
    cache_path = os.path.join(os.path.dirname(OUT_PATH), 'data_go_kr_kogas_supply.json')
    today = datetime.now(timezone.utc)

    # Stale cache check — refresh if missing or > 7 days old
    if os.path.exists(cache_path):
        try:
            with open(cache_path, 'r', encoding='utf-8') as f:
                cached = json.load(f)
            cached_at = cached.get('cachedAt')
            if cached_at:
                age = (today - datetime.fromisoformat(cached_at)).days
                if age < 7:
                    return cached
        except Exception:
            pass

    buf = _data_go_kr_download('15129906', '1')
    if not buf:
        return None

    text = _decode_korean_csv(buf)
    reader = _csv.DictReader(_io.StringIO(text))
    by_date = {}
    for r in reader:
        try:
            yr = int(r['연도']); mo = int(r['월'])
            use = r['용도'].strip()
            ton = int(str(r['공급량']).replace(',', ''))
        except (KeyError, ValueError):
            continue
        date = f'{yr}-{mo:02d}-01'
        if date not in by_date:
            by_date[date] = {'date': date}
        # 도시가스 = city gas, 발전용 = power generation
        key = ('cityGas_t' if '도시' in use else
                'power_t' if '발전' in use else use)
        by_date[date][key] = ton

    series = sorted(by_date.values(), key=lambda r: r['date'])
    # Compute total for convenience
    for r in series:
        r['total_t'] = (r.get('cityGas_t') or 0) + (r.get('power_t') or 0)

    out = {
        'cachedAt': today.isoformat(timespec='seconds'),
        'source':    'data.go.kr · KOGAS_용도별 월 공급량 (15129906)',
        'sourceUrl': f'{DATA_GO_KR_BASE}/data/15129906/fileData.do',
        'series':    series,
    }
    with open(cache_path, 'w', encoding='utf-8') as f:
        json.dump(out, f, indent=2, ensure_ascii=False)
    print(f'data.go.kr KOGAS supply: {len(series)} months '
          f'({series[0]["date"][:7]}–{series[-1]["date"][:7]})', file=sys.stderr)
    return out


def fetch_kogas_power_tariff():
    """Pull KOGAS power-sector natural gas tariff from Data.go.kr.
    No API key required. CSV columns:
      연월, 원(Nm3 원료비), 원(Nm3 공급비), 원(GJ 원료비), 원(GJ 공급비).
    Coverage: Jan 2008 → latest (irregular updates).

    Computes total tariff = raw material cost + supply cost (in both KRW/Nm³
    and KRW/GJ). USD/MMBtu conversion done client-side using ECB FX.
    """
    import csv as _csv
    import io as _io
    cache_path = os.path.join(os.path.dirname(OUT_PATH), 'data_go_kr_kogas_tariff.json')
    today = datetime.now(timezone.utc)
    if os.path.exists(cache_path):
        try:
            with open(cache_path, 'r', encoding='utf-8') as f:
                cached = json.load(f)
            cached_at = cached.get('cachedAt')
            if cached_at:
                age = (today - datetime.fromisoformat(cached_at)).days
                if age < 7:
                    return cached
        except Exception:
            pass

    buf = _data_go_kr_download('15052058', '1')
    if not buf:
        return None

    text = _decode_korean_csv(buf)
    reader = _csv.DictReader(_io.StringIO(text))
    series = []

    def _f(s):
        try:    return float(str(s).replace(',', ''))
        except (TypeError, ValueError): return None

    for r in reader:
        date = (r.get('연월') or '').strip()
        if not date:
            continue
        # Date may be 'YYYY-MM-DD' (most common) or 'YYYY-MM' or 'YYYYMM'
        if len(date) == 10 and date[4] == '-' and date[7] == '-':
            iso = date
        elif len(date) == 7 and date[4] == '-':
            iso = date + '-01'
        elif len(date) == 6:
            iso = f'{date[:4]}-{date[4:6]}-01'
        else:
            continue

        nm3_raw  = _f(r.get('원(Nm3 원료비)'))
        nm3_sup  = _f(r.get('원(Nm3 공급비)'))
        gj_raw   = _f(r.get('원(GJ 원료비)'))
        gj_sup   = _f(r.get('원(GJ 공급비)'))

        if nm3_raw is None and gj_raw is None:
            continue

        entry = {'date': iso}
        if nm3_raw is not None: entry['nm3_raw_krw']    = nm3_raw
        if nm3_sup is not None: entry['nm3_supply_krw'] = nm3_sup
        if gj_raw  is not None: entry['gj_raw_krw']     = gj_raw
        if gj_sup  is not None: entry['gj_supply_krw']  = gj_sup
        if gj_raw is not None and gj_sup is not None:
            entry['gj_total_krw'] = round(gj_raw + gj_sup, 2)
        if nm3_raw is not None and nm3_sup is not None:
            entry['nm3_total_krw'] = round(nm3_raw + nm3_sup, 2)
        series.append(entry)

    series.sort(key=lambda r: r['date'])
    out = {
        'cachedAt':  today.isoformat(timespec='seconds'),
        'source':    'data.go.kr · KOGAS_발전용 천연가스요금 (15052058)',
        'sourceUrl': f'{DATA_GO_KR_BASE}/data/15052058/fileData.do',
        'series':    series,
    }
    with open(cache_path, 'w', encoding='utf-8') as f:
        json.dump(out, f, indent=2, ensure_ascii=False)
    print(f'data.go.kr KOGAS tariff: {len(series)} months '
          f'({series[0]["date"][:7]}–{series[-1]["date"][:7]})', file=sys.stderr)
    return out


# ────────────────────────────────────────────────────────────────────
# UN Comtrade — Korea LNG imports by origin (HS 271111)
#
# Free account at https://comtradeplus.un.org/signin gives a subscription
# key via "Subscriptions" page. Free tier: 500 calls/day, 100k rows/call,
# monthly HS data, no Korean phone verification required.
#
# Endpoint:
#   https://comtradeapi.un.org/data/v1/get/C/M/HS
#   reporterCode=410 (Korea, Rep.)  flowCode=M (imports)  cmdCode=271111
#   period=YYYYMM     (or comma-separated YYYYMM list)
#   typeCode=C        (commodity)   freqCode=M (monthly)  clCode=HS
#
# Returns rows with: period (YYYYMM), partnerDesc (origin country name),
# netWgt (kg), primaryValue (USD CIF), partnerCode (M49 country code).
#
# We aggregate to: monthly Korea LNG imports total + per-origin breakdown.
# ────────────────────────────────────────────────────────────────────
COMTRADE_API_BASE = 'https://comtradeapi.un.org/data/v1/get/C/M/HS'


def fetch_comtrade_korea_lng_imports():
    """Pull Korea (410) monthly LNG imports HS 271111 from UN Comtrade.

    Requires COMTRADE_KEY env var (free key from comtradeplus.un.org).
    Returns None if key missing.

    Caches in data/comtrade_korea_lng.json keyed by period (YYYYMM).
    Only fetches missing periods + the latest 3 (in case of revisions).
    """
    api_key = os.environ.get('COMTRADE_KEY')
    if not api_key:
        print('COMTRADE_KEY not set — skipping UN Comtrade fetch', file=sys.stderr)
        return None

    cache_path = os.path.join(os.path.dirname(OUT_PATH), 'comtrade_korea_lng.json')
    cached = {'rowsByPeriod': {}}
    if os.path.exists(cache_path):
        try:
            with open(cache_path, 'r', encoding='utf-8') as f:
                cached = json.load(f)
        except Exception:
            cached = {'rowsByPeriod': {}}

    rows_by_period = cached.get('rowsByPeriod') or {}
    today = datetime.now(timezone.utc)

    # Build period list: 2018-01 through current month (Comtrade lags ~3 months)
    cutoff = today - timedelta(days=90)
    periods = []
    y, m = 2018, 1
    while (y, m) <= (cutoff.year, cutoff.month):
        periods.append(f'{y}{m:02d}')
        m += 1
        if m > 12: y += 1; m = 1

    # Refetch latest 6 (in case of revisions); skip already-cached others.
    refetch_recent = set(periods[-6:])
    todo = [p for p in periods
            if p not in rows_by_period or p in refetch_recent]

    print(f'Comtrade Korea LNG: {len(periods)} periods total, '
          f'{len(rows_by_period)} cached, {len(todo)} to fetch',
          file=sys.stderr)

    fetched = 0
    # Comtrade allows multi-period in one call (comma-separated, max ~12 at a time)
    BATCH = 12
    for i in range(0, len(todo), BATCH):
        chunk = todo[i:i + BATCH]
        url = (f'{COMTRADE_API_BASE}'
               f'?reporterCode=410&flowCode=M&cmdCode=271111'
               f'&period={",".join(chunk)}'
               f'&subscription-key={api_key}')
        try:
            text = fetch(url, timeout=60)
            data = json.loads(text)
        except Exception as e:
            print(f'Comtrade fetch failed (chunk {chunk[0]}…): {e}',
                  file=sys.stderr)
            time.sleep(2)
            continue
        items = data.get('data') or []
        if not items and data.get('error'):
            print(f"Comtrade error: {data.get('error')}", file=sys.stderr)
        for r in items:
            period = str(r.get('period') or '')
            partner = r.get('partnerDesc')
            net_wgt = r.get('netWgt')
            val = r.get('primaryValue')
            if not period or not partner:
                continue
            entry = rows_by_period.setdefault(period, {})
            entry[partner] = {
                'partnerCode': r.get('partnerCode'),
                'netWgt_kg':   net_wgt,
                'value_usd':   val,
            }
        for p in chunk:
            if p in rows_by_period:
                fetched += 1
        time.sleep(0.5)  # 500 calls/day — be polite

    # Build a flat monthly summary series
    series = []
    for period in sorted(rows_by_period.keys()):
        rows = rows_by_period[period]
        # Sum across all partners (excluding "World" if present)
        per_partner = {p: r for p, r in rows.items() if p != 'World'}
        total_kg = sum((r.get('netWgt_kg') or 0) for r in per_partner.values())
        total_usd = sum((r.get('value_usd') or 0) for r in per_partner.values())
        partners = sorted(per_partner.items(),
                          key=lambda kv: -(kv[1].get('netWgt_kg') or 0))
        top_partners = [{
            'partner':    name,
            'netWgt_kt':  round((r.get('netWgt_kg') or 0) / 1_000_000, 1),
            'value_musd': round((r.get('value_usd') or 0) / 1_000_000, 1),
        } for name, r in partners[:10]]
        series.append({
            'date':         f'{period[:4]}-{period[4:6]}-01',
            'period':       period,
            'totalImports_kt':    round(total_kg / 1_000_000, 1),
            'totalImports_musd':  round(total_usd / 1_000_000, 1),
            'cifPrice_usd_per_t': (round(total_usd / total_kg * 1000, 2)
                                    if total_kg > 0 else None),
            'topPartners':  top_partners,
        })

    out = {
        'cachedAt':       today.isoformat(timespec='seconds'),
        'source':         'UN Comtrade · Korea (410) imports HS 271111',
        'sourceUrl':      'https://comtradeplus.un.org/',
        'rowsByPeriod':   rows_by_period,
        'series':         series,
    }
    with open(cache_path, 'w', encoding='utf-8') as f:
        json.dump(out, f, indent=2, ensure_ascii=False)
    print(f'Comtrade Korea LNG: {len(series)} months in series '
          f'(fetched {fetched} new this run)', file=sys.stderr)
    return out


# ────────────────────────────────────────────────────────────────────
# KCGA — Korea City Gas Association monthly bulletin (XLSX)
#
# KCGA publishes a monthly Excel bulletin (도시가스사업통계월보) with the
# full sub-sector breakdown of city-gas demand that KOGAS does NOT
# disclose: Residential / General / Business / **Industrial** / CHP /
# Heat-only / Transport / Fuel-cell — by 34 distributors × 17 regions.
#
# Listing URL:  http://www.citygas.or.kr/info/monthly/index.jsp
# Bulletin URL: /info/monthly/read.jsp?reqPageNo=1&no=NNN  (no monotonic)
# Download URL: /include/download.jsp?path=...&vf=...&af=...
#
# Each XLSX is ~400KB. We cache by bulletin `no` so subsequent runs only
# fetch new bulletins.
# ────────────────────────────────────────────────────────────────────
KCGA_BASE = 'http://www.citygas.or.kr'

# Sub-sector total columns on Sheet 3 ('수요가수공급량(천㎥기준)') in the
# 전국 합계 row 34. Confirmed against Feb 2026 bulletin layout.
KCGA_VOLUME_COLS = {
    'residential_km3':       20,   # 가정용 소계 (cooking + heating)
    'general_km3':           23,   # 일반용 소계 (small commercial)
    'business_km3':          26,   # 업무용 소계 (large commercial / HVAC)
    'industrial_km3':        27,   # 산업용
    'chp_km3':               (28, 29),  # 열병합 1+2
    'heatOnly_km3':          30,   # 열전용 설비용
    'transport_km3':         31,   # 수송용 (CNG vehicles)
    'fuelCell_km3':          32,   # 연료전지용
    'total_km3':             33,   # 합 계
}

# LNG conversion: 1 m³ vapour ≈ 0.000737 t LNG (typical Korean spec)
KCGA_M3_PER_T = 1357.0  # ≈ 1/0.000737 — used to convert 천㎥ → kt for cross-check


def _kcga_list_bulletins():
    """Fetch the bulletin listing and return [{no, title, year, month, page}]."""
    import re
    rows = []
    # Pages: 7 pages of ~25 entries on the listing site.
    # Iterate until a page has no entries.
    for page in range(1, 12):
        url = f'{KCGA_BASE}/info/monthly/index.jsp?reqPageNo={page}'
        try:
            text = fetch(url, timeout=20)
        except Exception:
            break
        # Match anchors: <a href="read.jsp?reqPageNo=N&no=NNN">YYYY년 M월 도시가스사업통계월보</a>
        page_rows = re.findall(
            r'href="read\.jsp\?reqPageNo=\d+&no=(\d+)"[^>]*>\s*(?:<[^>]*>\s*)*([^<]*?)\s*(?:<|$)',
            text)
        added = 0
        for no, title in page_rows:
            title = title.strip()
            m = re.search(r'(\d{4})\s*년\s*(\d{1,2})\s*월', title)
            if m and '월보' in title:
                rows.append({
                    'no':    int(no),
                    'title': title,
                    'year':  int(m.group(1)),
                    'month': int(m.group(2)),
                    'page':  page,
                })
                added += 1
        if not added:
            break
    return rows


def _kcga_get_download_url(no):
    """Scrape the bulletin read page to extract the XLSX download URL."""
    import re
    page_url = f'{KCGA_BASE}/info/monthly/read.jsp?reqPageNo=1&no={no}'
    text = fetch(page_url, timeout=20)
    # Anchor like: href="/include/download.jsp?path=/upload/data/&vf=...&af=..."
    m = re.search(r'href="(/include/download\.jsp\?[^"]+)"', text)
    if not m:
        return None
    return KCGA_BASE + m.group(1)


def _kcga_parse_xlsx(buf):
    """Open an XLSX bulletin (in-memory bytes) and extract national totals
    + by-sub-sector volumes from sheet '3.수요가수공급량(천㎥기준)'."""
    try:
        import openpyxl
    except ImportError:
        print('openpyxl not installed — KCGA parse skipped', file=sys.stderr)
        return None
    import io as _io
    try:
        wb = openpyxl.load_workbook(_io.BytesIO(buf), data_only=True, read_only=True)
    except Exception as e:
        print(f'KCGA xlsx open failed: {e}', file=sys.stderr)
        return None

    target = '3.수요가수공급량(천㎥기준)'
    if target not in wb.sheetnames:
        # Some older bulletins may use slightly different naming
        candidates = [s for s in wb.sheetnames if '수요가수공급량' in s and '㎥' in s]
        if not candidates: return None
        target = candidates[0]
    ws = wb[target]

    # Find the 전국 합계 row by scanning column 1.
    nat_row = None
    for r in range(1, min(ws.max_row, 200) + 1):
        v = ws.cell(r, 1).value
        if v and ('전 국' in str(v) or '전국' in str(v)) and ('합' in str(v) or '계' in str(v)):
            nat_row = r
            break
    if not nat_row:
        return None

    out = {}
    for key, col in KCGA_VOLUME_COLS.items():
        if isinstance(col, tuple):
            v = sum((ws.cell(nat_row, c).value or 0) for c in col)
        else:
            v = ws.cell(nat_row, col).value
        if v is None:
            continue
        try:
            out[key] = round(float(v), 1)
        except (TypeError, ValueError):
            pass
    return out


def fetch_kcga_monthly():
    """Pull KCGA monthly bulletins → national city-gas demand by sub-sector.

    Caches in data/kcga_monthly.json by bulletin `no`. Each XLSX is ~400KB
    so initial backfill of ~150 months ~= 60MB downloads (do incrementally).
    Subsequent runs only fetch new bulletins.
    """
    cache_path = os.path.join(os.path.dirname(OUT_PATH), 'kcga_monthly.json')
    cached = {'series': [], 'parsedNos': []}
    if os.path.exists(cache_path):
        try:
            with open(cache_path, 'r', encoding='utf-8') as f:
                cached = json.load(f)
        except Exception:
            cached = {'series': [], 'parsedNos': []}

    parsed_set = set(cached.get('parsedNos') or [])
    series_by_date = {r['date']: r for r in (cached.get('series') or [])}

    try:
        bulletins = _kcga_list_bulletins()
    except Exception as e:
        print(f'KCGA list failed: {e}', file=sys.stderr)
        return None
    if not bulletins:
        print('KCGA: no bulletins listed', file=sys.stderr)
        return None

    # Per-run fetch budget — first run does heavy backfill, subsequent
    # runs catch only the latest 1-2 new ones. Cap to keep cron timing.
    PER_RUN_LIMIT = 25 if not parsed_set else 10
    todo = [b for b in sorted(bulletins, key=lambda x: x['no'], reverse=True)
            if b['no'] not in parsed_set][:PER_RUN_LIMIT]

    print(f'KCGA: {len(bulletins)} bulletins listed, '
          f'{len(parsed_set)} cached, {len(todo)} to fetch this run',
          file=sys.stderr)

    fetched = 0
    for b in todo:
        try:
            dl_url = _kcga_get_download_url(b['no'])
            if not dl_url:
                continue
            # Download XLSX (small ~400KB)
            if _USE_REQUESTS:
                r = _requests.get(dl_url, timeout=60,
                                  headers={'User-Agent': UA,
                                           'Referer': f'{KCGA_BASE}/info/monthly/read.jsp?no={b["no"]}'})
                r.raise_for_status()
                buf = r.content
            else:
                from urllib.request import urlopen, Request
                req = Request(dl_url, headers={'User-Agent': UA})
                buf = urlopen(req, timeout=60).read()
            parsed = _kcga_parse_xlsx(buf)
            if parsed:
                date = f'{b["year"]}-{b["month"]:02d}-01'
                row = {
                    'date':  date,
                    'no':    b['no'],
                    'title': b['title'],
                    **parsed,
                }
                series_by_date[date] = row
                parsed_set.add(b['no'])
                fetched += 1
        except Exception as e:
            print(f'KCGA bulletin no={b["no"]}: {e}', file=sys.stderr)

    series = sorted(series_by_date.values(), key=lambda r: r['date'])
    out = {
        'cachedAt':  datetime.now(timezone.utc).isoformat(timespec='seconds'),
        'source':    'Korea City Gas Association · 도시가스사업통계월보',
        'sourceUrl': f'{KCGA_BASE}/info/monthly/index.jsp',
        'parsedNos': sorted(parsed_set),
        'series':    series,
    }
    with open(cache_path, 'w', encoding='utf-8') as f:
        json.dump(out, f, indent=2, ensure_ascii=False)
    print(f'KCGA: {len(series)} months in series '
          f'(fetched {fetched} new this run)', file=sys.stderr)
    return out


def fetch_ecb_fx():
    """ECB daily reference rates (cross-rate to derive KRW/USD)."""
    url = 'https://www.ecb.europa.eu/stats/eurofxref/eurofxref-daily.xml'
    xml_text = fetch(url)
    root = ET.fromstring(xml_text)
    ns = {'ec': 'http://www.ecb.int/vocabulary/2002-08-01/eurofxref'}
    cube = root.find('.//ec:Cube[@time]', ns)
    if cube is None:
        raise RuntimeError('ECB feed missing time-stamped Cube')
    as_of = cube.get('time')
    rates = {c.get('currency'): float(c.get('rate'))
             for c in cube.findall('ec:Cube', ns)}
    eur_usd = rates.get('USD')
    eur_krw = rates.get('KRW')
    krw_usd = (eur_krw / eur_usd) if (eur_usd and eur_krw) else None
    return {
        'asOf':      as_of,
        'source':    'ECB euro reference rates',
        'sourceUrl': url,
        'eurUsd':    eur_usd,
        'eurKrw':    eur_krw,
        'krwUsd':    round(krw_usd, 2) if krw_usd else None,
    }


def round_for_diff_stability(out):
    """Round noisy numeric fields so hourly scrapes don't churn git when
    nothing material has changed. Reactor output drifting by 0.01% or 1 MWe
    isn't a meaningful event."""
    k = out.get('khnp') or {}
    for r in (k.get('reactors') or []):
        if r.get('output_pct') is not None:
            r['output_pct'] = round(r['output_pct'], 1)
        if r.get('output_mwe') is not None:
            r['output_mwe'] = round(r['output_mwe'])  # nearest integer MWe
    for s in (k.get('bySite') or []):
        s['totalCap_GW']  = round(s.get('totalCap_GW', 0), 2)
        s['onlineCap_GW'] = round(s.get('onlineCap_GW', 0), 2)
    if k.get('totalOnlineGW') is not None:
        k['totalOnlineGW'] = round(k['totalOnlineGW'], 2)


def append_history_row(out):
    """Append today's nuclear KPI to data/korea_nuclear_history.csv so we
    accumulate a current-year monthly utilization line over time. One row
    per scrape — duplicates per day are de-duped on read by the front-end."""
    k = out.get('khnp') or {}
    if not k or k.get('totalOnlineGW') is None:
        return
    history_path = os.path.join(os.path.dirname(OUT_PATH), 'korea_nuclear_history.csv')
    today = datetime.now(timezone.utc).strftime('%Y-%m-%d')
    # Compute aggregate util as online_GW / nominal_total (sum of all reactors'
    # nameplate). For a "true" utilization we'd need nameplate, not running output.
    # Approximation: sum of online_mwe + nameplate_for_offline. For now use
    # online_GW / nominal 25.0 GW (current Korean fleet nameplate) as the rough denominator.
    NOMINAL_FLEET_GW = 25.0
    util_pct = round(100 * k['totalOnlineGW'] / NOMINAL_FLEET_GW, 1)
    new_row = f"{today},{k['totalOnlineGW']},{NOMINAL_FLEET_GW},{util_pct},{k['onlineCount']},{k['totalCount']}\n"

    header = "date,online_GW,nominal_total_GW,util_pct,online_count,total_count\n"
    if not os.path.exists(history_path):
        with open(history_path, 'w', encoding='utf-8') as f:
            f.write(header)
            f.write(new_row)
        return
    # Skip if today's row already exists (overwrite with latest snapshot of the day)
    with open(history_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()
    today_idx = next((i for i, L in enumerate(lines) if L.startswith(today + ',')), None)
    if today_idx is not None:
        lines[today_idx] = new_row
    else:
        lines.append(new_row)
    with open(history_path, 'w', encoding='utf-8') as f:
        f.writelines(lines)


def main():
    out = {
        'updated':       datetime.now(timezone.utc).isoformat(timespec='seconds'),
        'fx':            None,
        'khnp':          None,
        'khnpAnnual':    None,
        'khnpTrips':     None,
        'emberMonthly':  None,
        # KR-only sources — left null so the front-end falls back to seeded values + amber badge.
        'kogasTariff':   None,
        'kogasImports':  None,
        # Gated on user OpenDART API key.
        'dart':          None,
        # Korea City Gas Association monthly bulletins — sub-sector demand
        'kcga':          None,
        'errors':        [],
    }

    try:
        out['fx'] = fetch_ecb_fx()
        print(f"FX: {out['fx']['krwUsd']} KRW/USD as of {out['fx']['asOf']}", file=sys.stderr)
    except Exception as e:
        out['errors'].append(f'ECB FX: {e}')
        print(f"ECB FX failed: {e}", file=sys.stderr)

    try:
        out['khnp'] = fetch_khnp_live()
        k = out['khnp']
        print(f"KHNP live: {k['onlineCount']}/{k['totalCount']} online · "
              f"{k['totalOnlineGW']} GW · {len(k['bySite'])} sites · "
              f"data as of {k.get('asOf')}", file=sys.stderr)
    except Exception as e:
        out['errors'].append(f'KHNP live: {e}')
        print(f"KHNP live failed: {e}", file=sys.stderr)

    try:
        out['khnpAnnual'] = fetch_khnp_annual_util()
        a = out['khnpAnnual']
        if a and a.get('byYear'):
            print(f"KHNP annual util: {len(a['byYear'])} years "
                  f"({a['byYear'][0]['year']}–{a['byYear'][-1]['year']})", file=sys.stderr)
    except Exception as e:
        out['errors'].append(f'KHNP annual util: {e}')
        print(f"KHNP annual util failed: {e}", file=sys.stderr)

    try:
        out['khnpTrips'] = fetch_khnp_trips()
        t = out['khnpTrips']
        if t and t.get('byYear'):
            print(f"KHNP unplanned trips: {len(t['byYear'])} years", file=sys.stderr)
    except Exception as e:
        out['errors'].append(f'KHNP trips: {e}')
        print(f"KHNP trips failed: {e}", file=sys.stderr)

    try:
        out['emberMonthly'] = fetch_ember_monthly()
        em = out['emberMonthly']
        if em and em.get('series'):
            print(f"Ember monthly: {len(em['series'])} months "
                  f"({em['series'][0]['date'][:7]}–{em['series'][-1]['date'][:7]})", file=sys.stderr)
    except Exception as e:
        out['errors'].append(f'Ember monthly: {e}')
        print(f"Ember monthly failed: {e}", file=sys.stderr)

    try:
        out['dart'] = fetch_dart_kogas()
        d = out['dart']
        if d:
            yrs = len(d.get('sectorVolumesByYear') or [])
            origins = d.get('lngOrigins') or []
            print(f"DART KOGAS: {d.get('latestReportName')} ({d.get('latestReportDate')}), "
                  f"{yrs} yrs sectoral volume, origins={origins}", file=sys.stderr)
    except Exception as e:
        out['errors'].append(f'DART KOGAS: {e}')
        print(f"DART KOGAS failed: {e}", file=sys.stderr)

    # Monthly KOGAS sales — separate fetcher (different filing type)
    try:
        monthly = fetch_dart_kogas_monthly()
        if monthly and out.get('dart') is not None:
            # Embed monthly series under dart.monthlySales so the front-end
            # only has to read one key for everything DART-derived.
            out['dart']['monthlySales'] = {
                'cachedAt':   monthly.get('cachedAt'),
                'source':     monthly.get('source'),
                'series':     monthly.get('series'),
            }
        if monthly:
            print(f"DART monthly: {len(monthly.get('series') or [])} months in series",
                  file=sys.stderr)
    except Exception as e:
        out['errors'].append(f'DART KOGAS monthly: {e}')
        print(f"DART KOGAS monthly failed: {e}", file=sys.stderr)

    # Quarterly inventory snapshots from KOGAS balance sheets via DART
    try:
        qtr = fetch_dart_kogas_quarterly_inventory()
        if qtr and out.get('dart') is not None:
            out['dart']['quarterlyInventory'] = {
                'cachedAt':   qtr.get('cachedAt'),
                'source':     qtr.get('source'),
                'snapshots':  qtr.get('snapshots'),
            }
        if qtr:
            print(f"DART quarterly inventory: {len(qtr.get('snapshots') or [])} snapshots",
                  file=sys.stderr)
    except Exception as e:
        out['errors'].append(f'DART quarterly inventory: {e}')
        print(f"DART quarterly inventory failed: {e}", file=sys.stderr)

    # data.go.kr KOGAS monthly supply by use (no API key)
    try:
        ds = fetch_kogas_monthly_supply()
        if ds:
            out['dataGoKr'] = out.get('dataGoKr') or {}
            out['dataGoKr']['kogasSupply'] = {
                'cachedAt': ds.get('cachedAt'),
                'source':   ds.get('source'),
                'series':   ds.get('series'),
            }
            print(f"data.go.kr KOGAS supply: {len(ds.get('series') or [])} months",
                  file=sys.stderr)
    except Exception as e:
        out['errors'].append(f'data.go.kr KOGAS supply: {e}')
        print(f"data.go.kr KOGAS supply failed: {e}", file=sys.stderr)

    # data.go.kr KOGAS power-sector tariff (no API key)
    try:
        dt = fetch_kogas_power_tariff()
        if dt:
            out['dataGoKr'] = out.get('dataGoKr') or {}
            out['dataGoKr']['kogasTariff'] = {
                'cachedAt': dt.get('cachedAt'),
                'source':   dt.get('source'),
                'series':   dt.get('series'),
            }
            print(f"data.go.kr KOGAS tariff: {len(dt.get('series') or [])} months",
                  file=sys.stderr)
    except Exception as e:
        out['errors'].append(f'data.go.kr KOGAS tariff: {e}')
        print(f"data.go.kr KOGAS tariff failed: {e}", file=sys.stderr)

    # Official KOGAS current-month power-sector tariff
    try:
        kt = fetch_kogas_current_power_tariff()
        if kt:
            out['kogasTariff'] = kt
            print(f"KOGAS current tariff: {kt.get('effectiveStart')}–{kt.get('effectiveEnd')}",
                  file=sys.stderr)
    except Exception as e:
        out['errors'].append(f'KOGAS current tariff: {e}')
        print(f"KOGAS current tariff failed: {e}", file=sys.stderr)

    # UN Comtrade — Korea LNG imports HS 271111 by origin (gated on COMTRADE_KEY)
    try:
        ct = fetch_comtrade_korea_lng_imports()
        if ct:
            out['comtrade'] = {
                'cachedAt': ct.get('cachedAt'),
                'source':   ct.get('source'),
                'series':   ct.get('series'),
            }
            print(f"Comtrade: {len(ct.get('series') or [])} months",
                  file=sys.stderr)
    except Exception as e:
        out['errors'].append(f'Comtrade: {e}')
        print(f"Comtrade failed: {e}", file=sys.stderr)

    # KCGA monthly bulletins — sub-sector demand (Res/Gen/Biz/Industrial/CHP/Transport)
    try:
        kcga = fetch_kcga_monthly()
        if kcga:
            out['kcga'] = {
                'cachedAt': kcga.get('cachedAt'),
                'source':   kcga.get('source'),
                'series':   kcga.get('series'),
            }
            print(f"KCGA: {len(kcga.get('series') or [])} months in series",
                  file=sys.stderr)
    except Exception as e:
        out['errors'].append(f'KCGA: {e}')
        print(f"KCGA failed: {e}", file=sys.stderr)

    # Stabilize for git diff
    round_for_diff_stability(out)

    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    with open(OUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(out, f, indent=2, ensure_ascii=False)
    print(f"Wrote {OUT_PATH}", file=sys.stderr)

    # History accumulator: append/update today's row in the CSV. The
    # workflow may pass --skip-history if it doesn't want this for some runs.
    if '--skip-history' not in sys.argv:
        try:
            append_history_row(out)
            print(f"Appended today's row to korea_nuclear_history.csv", file=sys.stderr)
        except Exception as e:
            out['errors'].append(f'history append: {e}')

    return 0 if not out['errors'] else 1


if __name__ == '__main__':
    sys.exit(main())
